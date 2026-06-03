import hashlib
import json
import os
import platform
import re
import shutil
import sys
import unicodedata
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import datetime
from pathlib import Path
from uuid import getnode

sys.dont_write_bytecode = True

import pandas as pd


# =========================
# CONFIGURACION DE RUTAS
# =========================
# Deja estas rutas vacias para usar las carpetas relativas del programa:
# entrada_xml, salida_excel, procesados y errores.
# Si algun dia necesitas una ruta personalizada, puedes escribirla aqui.
XML_FOLDER_MANUAL = ""

# Ruta opcional para PDF de ventas y retenciones emitidas.
# Vacia usa PDF\PDF VENTAS Y RETENCIONES EMITIDAS dentro del programa.
PDF_VENTAS_FOLDER_MANUAL = ""

# El programa lee una sola carpeta XML para compras, notas de credito y retenciones recibidas.
# Para PDF, por ahora solo lee ventas y retenciones emitidas desde PDF_VENTAS_FOLDER_MANUAL.
# Las carpetas PDF Compras, PDF Notas de Credito Recibidas y PDF Retenciones Recibidas
# se crean para que las uses con tus otros procesos.

# Ruta del Excel de salida.
# Si la dejas vacia, se crea "Consolidado XML.xlsx" junto a este Python.
# Si quieres controlar tambien la salida, escribe una ruta completa.
EXCEL_FILE_MANUAL = ""

# True lee XML dentro de subcarpetas. False solo lee los XML de la carpeta principal.
RECURSIVE = True

PDF_RECURSIVE = True

# PDF oficial del SRI usado para completar catalogos de ayuda en el Excel.
# Es opcional. Si no existe, el programa sigue funcionando.
ATS_FICHA_TECNICA_PDF_MANUAL = ""

# Si esta activo, mueve XML exitosos a procesados y XML con error a errores.
# En la app web se mantiene en False para trabajar solo con las carpetas configuradas.
MOVER_XML_PROCESADOS = False

# False deja los importes como numeros reales: 4.46
DECIMALES_CON_COMA = False

# =========================
# CONFIGURACION ATS
# =========================
# Valores del informante que iran en la cabecera del XML ATS.
# Usa "AUTO" para tomarlos desde los comprobantes.
ATS_TIPO_ID_INFORMANTE = "R"
ATS_ID_INFORMANTE = "AUTO"
ATS_RAZON_SOCIAL = "AUTO"
ATS_ANIO = "AUTO"
ATS_MES = "AUTO"
ATS_REGIMEN_MICROEMPRESA = ""
ATS_NUM_ESTAB_RUC = ""
ATS_TOTAL_VENTAS = 0.0
ATS_CODIGO_OPERATIVO = "IVA"

# Valores ATS por defecto para compras locales.
ATS_COD_SUSTENTO_DEFAULT = "02"
ATS_PARTE_REL_DEFAULT = "NO"
ATS_PAGO_LOC_EXT_DEFAULT = "01"
ATS_PAIS_EFEC_PAGO_DEFAULT = "NA"
ATS_APLIC_CONV_DOB_TRIB_DEFAULT = "NA"
ATS_PAG_EXT_SUJ_RET_NOR_LEG_DEFAULT = "NA"
ATS_TIPO_EMISION_VENTAS_PDF = "F"

# Si esta activo, ademas del Excel se crea el XML ATS junto al consolidado.
GENERAR_XML_ATS = False
ATS_XML_FILE_MANUAL = ""

# Evita exportar al XML documentos de otro mes/anio.
FILTRAR_ATS_POR_PERIODO = True

ATS_SUSTENTO_CATALOG = [
    ("01", "Credito Tributario para declaracion de IVA (servicios y bienes distintos de inventarios y activos fijos)", "01, 03, 04, 05, 11, 12, 21, 41, 43, 47, 48, 294, 344"),
    ("02", "Costo o Gasto para declaracion de IR (servicios y bienes distintos de inventarios y activos fijos)", "01, 02, 03, 04, 05, 09, 11, 12, 15, 19, 20, 21, 41, 43, 47, 48, 294, 344, 364"),
    ("03", "Activo Fijo - Credito Tributario para declaracion de IVA", "01, 03, 04, 05, 41, 47, 48, 294, 344"),
    ("04", "Activo Fijo - Costo o Gasto para declaracion de IR", "01, 02, 03, 04, 05, 15, 41, 47, 48, 294, 344"),
    ("05", "Liquidacion Gastos de Viaje, hospedaje y alimentacion Gastos IR", "01, 02, 03, 04, 05, 11, 15, 41, 294, 344"),
    ("06", "Inventario - Credito Tributario para declaracion de IVA", "01, 03, 04, 05, 41, 43, 47, 48, 294, 344"),
    ("07", "Inventario - Costo o Gasto para declaracion de IR", "01, 02, 03, 04, 05, 15, 41, 43, 47, 48, 294, 344, 364"),
    ("08", "Valor pagado para solicitar Reembolso de Gasto (intermediario)", "01, 02, 03, 04, 05, 21, 294, 344"),
    ("09", "Reembolso por Siniestros", "01, 04, 05, 45"),
    ("10", "Distribucion de Dividendos, Beneficios o Utilidades", "19"),
    ("11", "Convenios de debito o recaudacion para IFI", "12"),
    ("12", "Impuestos y retenciones presuntivos", "42"),
    ("13", "Valores reconocidos por entidades del sector publico a favor de sujetos pasivos", "19"),
    ("14", "Valores facturados por socios a operadoras de transporte", "01, 02, 03, 04, 05"),
    ("15", "Pagos efectuados por consumos propios y de terceros de servicios digitales", "01, 02, 03, 04, 05, 12, 15"),
    ("00", "Casos especiales cuyo sustento no aplica en las opciones anteriores (no vigente desde 28/02/2015)", "01, 02, 04, 05, 19, 42"),
]

ATS_TIPO_COMPROBANTE_CATALOG = [
    ("01", "Factura"),
    ("02", "Nota o boleta de venta"),
    ("03", "Liquidacion de compra de bienes o prestacion de servicios"),
    ("04", "Nota de credito"),
    ("05", "Nota de debito"),
    ("06", "Guias de remision"),
    ("07", "Comprobante de retencion"),
    ("08", "Boletos o entradas a espectaculos publicos"),
    ("09", "Tiquetes o vales emitidos por maquinas registradoras"),
    ("11", "Pasajes expedidos por empresas de aviacion"),
    ("12", "Documentos emitidos por instituciones financieras"),
    ("15", "Comprobante de venta emitido en el exterior"),
    ("16", "Formulario Unico de Exportacion / DAU / DAV"),
    ("18", "Documentos autorizados utilizados en ventas excepto N/C N/D"),
    ("19", "Comprobantes de pago de cuotas o aportes"),
    ("20", "Documentos por servicios administrativos del Estado"),
    ("21", "Carta de Porte Aereo"),
    ("22", "RECAP"),
    ("23", "Nota de Credito TC"),
    ("24", "Nota de Debito TC"),
    ("41", "Comprobante de venta emitido por reembolso"),
    ("42", "Documento retencion presuntiva y retencion emitida por propio vendedor o intermediario"),
    ("43", "Liquidacion para explotacion y exploracion de hidrocarburos"),
    ("44", "Comprobante de contribuciones y aportes"),
    ("45", "Liquidacion por reclamos de aseguradoras"),
    ("47", "Nota de credito por reembolso emitida por intermediario"),
    ("48", "Nota de debito por reembolso emitida por intermediario"),
    ("49", "Proveedor directo de exportador bajo regimen especial"),
    ("50", "A instituciones del Estado y empresas publicas que perciben ingreso exento"),
    ("51", "N/C a instituciones del Estado y empresas publicas que perciben ingreso exento"),
    ("52", "N/D a instituciones del Estado y empresas publicas que perciben ingreso exento"),
    ("294", "Liquidacion de compra de bienes muebles usados"),
    ("344", "Liquidacion de compra de vehiculos usados"),
    ("364", "Acta Entrega-Recepcion PET"),
    ("370", "Factura operadora transporte / socio"),
    ("371", "Comprobante socio a operadora de transporte"),
    ("372", "Nota de credito operadora transporte / socio"),
    ("373", "Nota de debito operadora transporte / socio"),
    ("374", "Nota de debito operadora transporte / socio"),
    ("375", "Liquidacion de compra RISE de bienes o prestacion de servicios"),
]

ATS_FORMA_PAGO_CATALOG = [
    ("01", "Sin utilizacion del sistema financiero", "01/01/2013", "-"),
    ("02", "Cheque propio", "01/01/2013", "31/08/2016"),
    ("03", "Cheque certificado", "01/01/2013", "31/08/2016"),
    ("04", "Cheque de gerencia", "01/01/2013", "31/08/2016"),
    ("05", "Cheque del exterior", "01/01/2013", "31/08/2016"),
    ("06", "Debito de cuenta", "01/01/2013", "31/08/2016"),
    ("07", "Transferencia propio banco", "01/01/2013", "31/08/2016"),
    ("08", "Transferencia otro banco nacional", "01/01/2013", "31/08/2016"),
    ("09", "Transferencia banco exterior", "01/01/2013", "31/08/2016"),
    ("10", "Tarjeta de credito nacional", "01/01/2013", "31/08/2016"),
    ("11", "Tarjeta de credito internacional", "01/01/2013", "31/08/2016"),
    ("12", "Giro", "01/01/2013", "31/08/2016"),
    ("13", "Deposito en cuenta corriente/ahorros", "01/01/2013", "31/08/2016"),
    ("14", "Endoso de inversion", "01/01/2013", "31/08/2016"),
    ("15", "Compensacion de deudas", "01/01/2013", "-"),
    ("16", "Tarjeta de debito", "01/05/2016", "-"),
    ("17", "Dinero electronico", "01/05/2016", "-"),
    ("18", "Tarjeta prepago", "01/05/2016", "-"),
    ("19", "Tarjeta de credito", "01/06/2016", "-"),
    ("20", "Otros con utilizacion del sistema financiero", "01/06/2016", "-"),
    ("21", "Endoso de titulos", "01/06/2016", "-"),
]

ATS_RETENCION_IVA_CATALOG = [
    ("9", "10", "01/06/2015", "-"),
    ("10", "20", "01/06/2015", "-"),
    ("1", "30", "01/01/2002", "-"),
    ("11", "50", "01/01/2016", "-"),
    ("2", "70", "01/01/2002", "-"),
    ("3", "100", "01/01/2002", "-"),
]

_RETENCION_RENTA_CATALOG_CACHE = None

ATS_TIPO_PAGO_CATALOG = [
    ("01", "Pago a residente / establecimiento permanente"),
    ("02", "Pago a no residente"),
]

ATS_SI_NO_NA_CATALOG = [
    ("NA", "No aplica"),
    ("SI", "Si"),
    ("NO", "No"),
]

SCRIPT_DIR = Path(__file__).resolve().parent
if SCRIPT_DIR.name.upper() == "EJECUTOR PYTHON" and SCRIPT_DIR.parent.name.upper() == "CONFIGURACION":
    BASE_DIR = SCRIPT_DIR.parent.parent
else:
    BASE_DIR = SCRIPT_DIR
CONFIG_DIR = BASE_DIR / "CONFIGURACION"
ATS_DIR = BASE_DIR / "ATS"
PDF_ROOT_DIR = BASE_DIR / "PDF"
SALIDA_EXCEL_DIR = BASE_DIR
PROCESADOS_DIR = BASE_DIR / "procesados"
ERRORES_DIR = BASE_DIR / "errores"
XML_FOLDER = Path(XML_FOLDER_MANUAL) if XML_FOLDER_MANUAL.strip() else BASE_DIR / "XML"
PDF_COMPRAS_FOLDER = PDF_ROOT_DIR / "PDF Compras"
PDF_NC_RECIBIDAS_FOLDER = PDF_ROOT_DIR / "PDF Notas de Credito Recibidas"
PDF_RET_RECIBIDAS_FOLDER = PDF_ROOT_DIR / "PDF Retenciones Recibidas"
PDF_VENTAS_FOLDER = Path(PDF_VENTAS_FOLDER_MANUAL) if PDF_VENTAS_FOLDER_MANUAL else PDF_ROOT_DIR / "PDF VENTAS Y RETENCIONES EMITIDAS"
EXCEL_FILE = Path(EXCEL_FILE_MANUAL) if EXCEL_FILE_MANUAL else SALIDA_EXCEL_DIR / "Consolidado XML.xlsx"

# =========================
# ACTIVACION LOCAL
# =========================
# Pide clave solo la primera vez en cada computadora.
# La clave no queda escrita en texto, solo se compara contra este hash.
REQUIERE_ACTIVACION = True
ACTIVACION_FILE = CONFIG_DIR / ".activacion_consolidado"
ACTIVACION_SALT = "CONSOLIDADO_XML_ACTIVACION_V1"
CLAVE_ACTIVACION_HASH_SHA256 = "96c051773aea19ef1d8b167d031b0b3aac6e45678bf4bf9881dc70978cedea68"

COMPRAS_COLUMNS = [
    "Razon Social Comprador",
    "Identificacion Comprador",
    "Razon Social Emisor",
    "RUC Emisor",
    "Fecha Emision",
    "Fecha Autorizacion",
    "Numero de Factura",
    "Subtotal IVA 0%",
    "IVA 0%",
    "Subtotal IVA 5%",
    "IVA 5%",
    "Subtotal IVA 8%",
    "IVA 8%",
    "Subtotal IVA 12%",
    "IVA 12%",
    "Subtotal IVA 14%",
    "IVA 14%",
    "Subtotal IVA 15%",
    "IVA 15%",
    "No Objeto IVA",
    "Exento IVA",
    "Total Sin Impuestos",
    "Total",
    "Producto 1",
    "Producto 2",
    "Producto 3",
    "Numero Autorizacion",
]

COMPRAS_NUMERIC_COLS = [
    "Subtotal IVA 0%", "IVA 0%",
    "Subtotal IVA 5%", "IVA 5%",
    "Subtotal IVA 8%", "IVA 8%",
    "Subtotal IVA 12%", "IVA 12%",
    "Subtotal IVA 14%", "IVA 14%",
    "Subtotal IVA 15%", "IVA 15%",
    "No Objeto IVA", "Exento IVA",
    "Total Sin Impuestos", "Total",
]

NC_COLUMNS = [
    "Razon Social Comprador",
    "Identificacion Comprador",
    "Razon Social Emisor",
    "RUC Emisor",
    "Fecha Emision",
    "Numero de Nota de Crédito",
    "Numero Factura Modificada",
    "Subtotal IVA 0%",
    "IVA 0%",
    "Subtotal IVA 5%",
    "IVA 5%",
    "Subtotal IVA 8%",
    "IVA 8%",
    "Subtotal IVA 12%",
    "IVA 12%",
    "Subtotal IVA 14%",
    "IVA 14%",
    "Subtotal IVA 15%",
    "IVA 15%",
    "Subtotal No Objeto",
    "No Objeto",
    "Subtotal Exento",
    "Total",
    "Numero Autorización",
    "Detalle 1",
    "Detalle 2",
    "Detalle 3",
]

NC_NUMERIC_COLS = [
    "Subtotal IVA 0%", "IVA 0%",
    "Subtotal IVA 5%", "IVA 5%",
    "Subtotal IVA 8%", "IVA 8%",
    "Subtotal IVA 12%", "IVA 12%",
    "Subtotal IVA 14%", "IVA 14%",
    "Subtotal IVA 15%", "IVA 15%",
    "Subtotal No Objeto", "No Objeto",
    "Subtotal Exento", "Total",
]

RET_COLUMNS = [
    "Razon Social Emisor",
    "RUC Emisor",
    "Establecimiento",
    "Punto Emision",
    "Numero Comprobante Retencion",
    "Razon Social Sujeto Retenido",
    "Identificacion Sujeto Retenido",
    "Fecha Emision",
    "Periodo Fiscal",
    "Numero Doc Sustento",
    "Fecha Doc Sustento",
    "Total sin Impuestos",
    "Importe Total",
    "Tipo Retencion",
    "Codigo Retencion",
    "Base Imponible",
    "Porcentaje Retencion",
    "Valor Retenido",
    "Clave Acceso",
    "Numero Autorizacion",
]

RET_NUMERIC_COLS = [
    "Total sin Impuestos",
    "Importe Total",
    "Base Imponible",
    "Porcentaje Retencion",
    "Valor Retenido",
]

VENTAS_COLUMNS = [
    "Fuente PDF",
    "Exportar XML",
    "Estado ATS",
    "Observacion",
    "Razon Social Emisor",
    "RUC Emisor",
    "Razon Social Cliente",
    "Identificacion Cliente",
    "Fecha Emision",
    "Fecha Autorizacion",
    "Numero de Factura",
    "Establecimiento",
    "Punto Emision",
    "Secuencial",
    "Numero Autorizacion",
    "Subtotal IVA 0%",
    "IVA 0%",
    "Subtotal IVA 5%",
    "Subtotal IVA 8%",
    "Subtotal IVA 12%",
    "Subtotal IVA 14%",
    "Subtotal IVA 15%",
    "IVA 5%",
    "IVA 8%",
    "IVA 12%",
    "IVA 14%",
    "IVA 15%",
    "No Objeto IVA",
    "Exento IVA",
    "ICE",
    "Total Sin Impuestos",
    "Total",
    "Forma Pago",
    "Producto 1",
    "Producto 2",
    "Producto 3",
    "tpIdCliente",
    "idCliente",
    "parteRelVtas",
    "tipoComprobante",
    "tipoEmision",
    "numeroComprobantes",
    "valorRetIva",
    "valorRetRenta",
]

VENTAS_NUMERIC_COLS = [
    "Subtotal IVA 0%",
    "IVA 0%",
    "Subtotal IVA 5%",
    "Subtotal IVA 8%",
    "Subtotal IVA 12%",
    "Subtotal IVA 14%",
    "Subtotal IVA 15%",
    "IVA 5%",
    "IVA 8%",
    "IVA 12%",
    "IVA 14%",
    "IVA 15%",
    "No Objeto IVA",
    "Exento IVA",
    "ICE",
    "Total Sin Impuestos",
    "Total",
    "numeroComprobantes",
    "valorRetIva",
    "valorRetRenta",
]

VENTAS_SUMMARY_COLS = [
    "Subtotal IVA 0%",
    "IVA 0%",
    "Subtotal IVA 5%",
    "IVA 5%",
    "Subtotal IVA 8%",
    "IVA 8%",
    "Subtotal IVA 12%",
    "IVA 12%",
    "Subtotal IVA 14%",
    "IVA 14%",
    "Subtotal IVA 15%",
    "IVA 15%",
    "No Objeto IVA",
    "Exento IVA",
    "ICE",
    "Total Sin Impuestos",
    "Total",
    "valorRetIva",
    "valorRetRenta",
]

VENTAS_DATE_COLS = {"Fecha Emision", "Fecha Autorizacion"}

RET_EMITIDAS_COLUMNS = [
    "Fuente PDF",
    "Estado",
    "Observacion",
    "Razon Social Emisor",
    "RUC Emisor",
    "Numero Comprobante Retencion",
    "Razon Social Sujeto Retenido",
    "Identificacion Sujeto Retenido",
    "Fecha Emision",
    "Periodo Fiscal",
    "Numero Doc Sustento",
    "Fecha Doc Sustento",
    "Tipo Retencion",
    "Codigo Retencion",
    "Base Imponible",
    "Porcentaje Retencion",
    "Valor Retenido",
    "Numero Autorizacion",
]

RET_EMITIDAS_NUMERIC_COLS = [
    "Base Imponible",
    "Porcentaje Retencion",
    "Valor Retenido",
]

RET_EMITIDAS_DATE_COLS = {"Fecha Emision", "Fecha Doc Sustento"}

ALERTAS_COLUMNS = [
    "Tipo",
    "Detalle",
    "Fuente",
]

VENTAS_CODE_TEXT_COLS = {
    "RUC Emisor",
    "Identificacion Cliente",
    "Numero de Factura",
    "Establecimiento",
    "Punto Emision",
    "Secuencial",
    "Numero Autorizacion",
    "Forma Pago",
    "tpIdCliente",
    "idCliente",
    "parteRelVtas",
    "tipoComprobante",
    "tipoEmision",
}

VENTAS_XML_NUMERIC_COLS = {
    "baseNoGraIva",
    "baseImponible",
    "baseImpGrav",
    "montoIva",
    "montoIce",
    "valorRetIva",
    "valorRetRenta",
}

ATS_META_COLUMNS = [
    "Exportar XML",
    "Estado ATS",
    "Observacion",
    "Fuente",
    "TipoIDInformante",
    "IdInformante",
    "razonSocial",
    "Anio",
    "Mes",
    "regimenMicroempresa",
    "numEstabRuc",
    "totalVentas",
    "codigoOperativo",
]

ATS_DETAIL_COLUMNS = [
    "codSustento",
    "tpIdProv",
    "idProv",
    "tipoComprobante",
    "tipoProv",
    "denoProv",
    "parteRel",
    "fechaRegistro",
    "establecimiento",
    "puntoEmision",
    "secuencial",
    "fechaEmision",
    "autorizacion",
    "baseNoGraIva",
    "baseImponible",
    "baseImpGrav",
    "baseImpExe",
    "montoIce",
    "montoIva",
    "valRetBien10",
    "valRetServ20",
    "valorRetBienes",
    "valRetServ50",
    "valorRetServicios",
    "valRetServ100",
    "valorRetencionNc",
    "totbasesImpReemb",
    "pagoLocExt",
    "paisEfecPago",
    "aplicConvDobTrib",
    "pagExtSujRetNorLeg",
    "formaPago",
    "codRetAir",
    "baseImpAir",
    "porcentajeAir",
    "valRetAir",
    "air_detalle",
    "estabRetencion1",
    "ptoEmiRetencion1",
    "secRetencion1",
    "autRetencion1",
    "fechaEmiRet1",
    "docModificado",
    "estabModificado",
    "ptoEmiModificado",
    "secModificado",
    "autModificado",
]

ATS_COLUMNS = ATS_META_COLUMNS + ATS_DETAIL_COLUMNS

ATS_NUMERIC_COLS = [
    "totalVentas",
    "baseNoGraIva",
    "baseImponible",
    "baseImpGrav",
    "baseImpExe",
    "montoIce",
    "montoIva",
    "valRetBien10",
    "valRetServ20",
    "valorRetBienes",
    "valRetServ50",
    "valorRetServicios",
    "valRetServ100",
    "valorRetencionNc",
    "totbasesImpReemb",
    "baseImpAir",
    "porcentajeAir",
    "valRetAir",
]

ATS_DATE_COLS = {"fechaRegistro", "fechaEmision", "fechaEmiRet1"}

ATS_DETAIL_XML_ORDER = [
    "codSustento",
    "tpIdProv",
    "idProv",
    "tipoComprobante",
    "tipoProv",
    "denoProv",
    "parteRel",
    "fechaRegistro",
    "establecimiento",
    "puntoEmision",
    "secuencial",
    "fechaEmision",
    "autorizacion",
    "baseNoGraIva",
    "baseImponible",
    "baseImpGrav",
    "baseImpExe",
    "montoIce",
    "montoIva",
    "valRetBien10",
    "valRetServ20",
    "valorRetBienes",
    "valRetServ50",
    "valorRetServicios",
    "valRetServ100",
    "valorRetencionNc",
    "totbasesImpReemb",
]

ATS_REQUIRED_XML_FIELDS = {
    "codSustento",
    "tpIdProv",
    "idProv",
    "tipoComprobante",
    "parteRel",
    "fechaRegistro",
    "establecimiento",
    "puntoEmision",
    "secuencial",
    "fechaEmision",
    "autorizacion",
    "baseNoGraIva",
    "baseImponible",
    "baseImpGrav",
    "baseImpExe",
    "montoIce",
    "montoIva",
    "valorRetBienes",
    "valorRetServicios",
    "valRetServ100",
}

ATS_OPTIONAL_XML_FIELDS = {
    "tipoProv",
    "denoProv",
    "valRetBien10",
    "valRetServ20",
    "valRetServ50",
    "valorRetencionNc",
    "totbasesImpReemb",
}

ATS_CODE_TEXT_COLS = {
    "TipoIDInformante",
    "IdInformante",
    "Anio",
    "Mes",
    "codigoOperativo",
    "codSustento",
    "tpIdProv",
    "idProv",
    "tipoComprobante",
    "tipoProv",
    "parteRel",
    "establecimiento",
    "puntoEmision",
    "secuencial",
    "autorizacion",
    "pagoLocExt",
    "paisEfecPago",
    "aplicConvDobTrib",
    "pagExtSujRetNorLeg",
    "formaPago",
    "codRetAir",
    "estabRetencion1",
    "ptoEmiRetencion1",
    "secRetencion1",
    "autRetencion1",
    "docModificado",
    "estabModificado",
    "ptoEmiModificado",
    "secModificado",
    "autModificado",
}

ATS_EDITABLE_COLUMNS = {
    "codSustento",
    "tipoComprobante",
    "parteRel",
    "tpIdProv",
    "pagoLocExt",
    "paisEfecPago",
    "aplicConvDobTrib",
    "pagExtSujRetNorLeg",
    "formaPago",
    "codRetAir",
    "baseImpAir",
    "porcentajeAir",
    "valRetAir",
    "docModificado",
}

IVA_CODPORC_TO_RATE = {
    "0": 0.0,
    "5": 5.0,
    "8": 8.0,
    "2": 12.0,
    "3": 14.0,
    "4": 15.0,
}

CODPORC_NO_OBJETO = {"6"}
CODPORC_EXENTO = {"7"}
TIPO_RETENCION_MAP = {"1": "RENTA", "2": "IVA", "6": "ISD"}


def calcular_hash_script():
    return hashlib.sha256(Path(__file__).read_bytes()).hexdigest()


def obtener_id_maquina():
    return f"{platform.node()}|{getnode()}"


def calcular_token_activacion(script_hash):
    raw = f"{script_hash}|{obtener_id_maquina()}|{ACTIVACION_SALT}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def leer_activacion():
    try:
        return json.loads(ACTIVACION_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def guardar_activacion(script_hash):
    ACTIVACION_FILE.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "script_hash": script_hash,
        "machine": obtener_id_maquina(),
        "token": calcular_token_activacion(script_hash),
    }
    ACTIVACION_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")


def activacion_valida(script_hash):
    data = leer_activacion()
    return (
        data.get("script_hash") == script_hash
        and data.get("token") == calcular_token_activacion(script_hash)
    )


def pedir_clave_activacion():
    print("Clave de instalacion: ", end="", flush=True)
    clave = sys.stdin.readline()
    clave = clave.strip().strip('"').strip("'")
    return clave.replace("\ufeff", "").replace("\u200b", "")


def verificar_activacion():
    if not REQUIERE_ACTIVACION:
        return

    script_hash = calcular_hash_script()
    if activacion_valida(script_hash):
        return

    clave = pedir_clave_activacion()
    clave_hash = hashlib.sha256(clave.encode("utf-8")).hexdigest()
    if clave_hash != CLAVE_ACTIVACION_HASH_SHA256:
        print("Clave incorrecta. No se genero ningun archivo.")
        sys.exit(1)

    guardar_activacion(script_hash)
    print("Activacion correcta. Esta computadora quedo autorizada para esta version.")


def strip_namespaces(elem):
    for item in elem.iter():
        if isinstance(item.tag, str) and "}" in item.tag:
            item.tag = item.tag.split("}", 1)[1]
    return elem


def get_text(node, path, default=""):
    if node is None:
        return default
    found = node.find(path)
    if found is None or found.text is None:
        return default
    return found.text.strip()


def to_float(value, default=0.0):
    try:
        if value is None:
            return default
        text = str(value).strip()
        if not text:
            return default
        return float(text.replace(",", "."))
    except Exception:
        return default


def parse_date_any(value):
    if not value:
        return None
    value = value.strip()
    if re.match(r".*[+-]\d{2}:\d{2}$", value):
        value = value[:-3] + value[-2:]

    for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except Exception:
            pass
    return None


def safe_date_iso(raw):
    parsed = parse_date_any(raw)
    return parsed.date().isoformat() if parsed else (raw or "")


def list_xml_files(folder, recursive=True):
    folder = Path(folder)
    if not folder.exists():
        folder.mkdir(parents=True, exist_ok=True)
        return []
    pattern = "**/*.xml" if recursive else "*.xml"
    return sorted(folder.glob(pattern))


def list_pdf_files(folder, recursive=True):
    folder = Path(folder)
    if not folder.exists():
        folder.mkdir(parents=True, exist_ok=True)
        return []
    pattern = "**/*.pdf" if recursive else "*.pdf"
    return sorted(folder.glob(pattern))


def format_decimal_comma(value):
    try:
        if value is None:
            return ""
        if isinstance(value, float) and pd.isna(value):
            return ""
        return f"{float(value):.2f}".replace(".", ",")
    except Exception:
        return "" if value is None else str(value)


def clean_text(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_digits(value):
    text = clean_text(value)
    if re.match(r"^\d+\.0$", text):
        text = text[:-2]
    return re.sub(r"\D", "", text)


def clean_integer_text(value, strip_leading=True):
    digits = clean_digits(value)
    if not digits:
        return ""
    if strip_leading:
        digits = digits.lstrip("0") or "0"
    return digits


def clean_fixed_digits(value, width):
    digits = clean_digits(value)
    if not digits:
        return ""
    return digits.zfill(width)[-width:]


def xml_decimal(value):
    return f"{to_float(value):.2f}"


def date_ddmmyyyy(value):
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = clean_text(value)
    if not text:
        return ""
    parsed = parse_date_any(text)
    return parsed.strftime("%d/%m/%Y") if parsed else text


def split_document_number(value):
    text = clean_text(value)
    parts = re.findall(r"\d+", text)
    if len(parts) >= 3:
        return (
            parts[0].zfill(3)[-3:],
            parts[1].zfill(3)[-3:],
            (parts[2].lstrip("0") or "0"),
        )

    digits = clean_digits(text)
    if len(digits) >= 15:
        return digits[:3], digits[3:6], (digits[6:].lstrip("0") or "0")
    return "", "", ""


def document_match_key(value):
    return clean_digits(value)


def tipo_id_from_identificacion(value):
    digits = clean_digits(value)
    if len(digits) == 13:
        return "01"
    if len(digits) == 10:
        return "02"
    return "03" if digits else ""


def tipo_id_cliente_from_identificacion(value):
    digits = clean_digits(value)
    if digits == "9999999999999":
        return "07"
    if len(digits) == 13:
        return "04"
    if len(digits) == 10:
        return "05"
    return "06" if digits else ""


def sum_columns(row, columns):
    return sum(to_float(row.get(column, 0)) for column in columns)


def normalize_header(value):
    text = unicodedata.normalize("NFKD", clean_text(value))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def first_existing(row, names, default=""):
    wanted = {normalize_header(name) for name in names}
    for name in getattr(row, "index", []):
        if normalize_header(name) in wanted and clean_text(row.get(name)):
            return row.get(name)
    return default


def is_auto(value):
    return clean_text(value).upper() in {"", "AUTO"}


def configured_or_auto(config_value, inferred_value):
    return clean_text(inferred_value) if is_auto(config_value) else clean_text(config_value)


def most_common_period(values):
    periods = []
    for value in values:
        parsed = parse_date_any(clean_text(value))
        if parsed:
            periods.append((parsed.year, parsed.month))
    if not periods:
        return "", ""
    (year, month), _ = Counter(periods).most_common(1)[0]
    return str(year), f"{month:02d}"


def first_df_value(df, column):
    if df is None or df.empty or column not in df.columns:
        return ""
    for value in df[column]:
        if clean_text(value):
            return clean_text(value)
    return ""


def contribuyente_key(value):
    digits = clean_digits(value)
    if len(digits) == 13 and digits.endswith("001"):
        return digits[:10]
    return digits


def contribuyente_display_id(value):
    return clean_digits(value)


def most_common_nonempty(values):
    cleaned = [clean_text(value) for value in values if clean_text(value)]
    if not cleaned:
        return ""
    return Counter(cleaned).most_common(1)[0][0]


def infer_contribuyente(compras_df=None, nc_df=None, ret_df=None, ventas_df=None):
    candidates = []
    for df, id_col, name_col in (
        (compras_df, "Identificacion Comprador", "Razon Social Comprador"),
        (nc_df, "Identificacion Comprador", "Razon Social Comprador"),
        (ret_df, "Identificacion Sujeto Retenido", "Razon Social Sujeto Retenido"),
        (ventas_df, "RUC Emisor", "Razon Social Emisor"),
    ):
        if df is None or df.empty or id_col not in df.columns:
            continue
        for _, row in df.iterrows():
            key = contribuyente_key(row.get(id_col))
            if key:
                candidates.append((
                    key,
                    contribuyente_display_id(row.get(id_col)),
                    clean_text(row.get(name_col)) if name_col in df.columns else "",
                ))

    if not candidates:
        return {"key": "", "id": "", "razon": ""}

    key = Counter(item[0] for item in candidates).most_common(1)[0][0]
    ids = [item[1] for item in candidates if item[0] == key and item[1]]
    names = [item[2] for item in candidates if item[0] == key and item[2]]
    preferred_ids = sorted(ids, key=lambda item: (len(item), item), reverse=True)
    return {
        "key": key,
        "id": preferred_ids[0] if preferred_ids else key,
        "razon": most_common_nonempty(names),
    }


def add_contribuyente_alerts(alerts, df, id_col, name_col, expected, label):
    if df is None or df.empty or id_col not in df.columns or not expected.get("key"):
        return

    for _, row in df.iterrows():
        value = row.get(id_col)
        key = contribuyente_key(value)
        if key and key != expected["key"]:
            source = (
                row.get("Fuente PDF")
                or row.get("Numero de Factura")
                or row.get("Numero de Nota de Crédito")
                or row.get("Numero Comprobante Retencion")
                or ""
            )
            alerts.append({
                "Tipo": "NO COINCIDE CONTRIBUYENTE",
                "Detalle": (
                    f"{label}: {clean_text(row.get(name_col))} / {clean_text(value)} "
                    f"no coincide con {expected.get('razon')} / {expected.get('id')}"
                ),
                "Fuente": clean_text(source),
            })


def build_alertas_dataframe(compras_df, nc_df, ret_df, ventas_df, ret_emitidas_df, contribuyente):
    alerts = []
    if not contribuyente.get("key"):
        alerts.append({
            "Tipo": "SIN CONTRIBUYENTE",
            "Detalle": "No se pudo reconocer un contribuyente principal desde compras, notas, retenciones o ventas.",
            "Fuente": "",
        })

    add_contribuyente_alerts(alerts, compras_df, "Identificacion Comprador", "Razon Social Comprador", contribuyente, "Compras")
    add_contribuyente_alerts(alerts, nc_df, "Identificacion Comprador", "Razon Social Comprador", contribuyente, "Notas de credito")
    add_contribuyente_alerts(alerts, ret_df, "Identificacion Sujeto Retenido", "Razon Social Sujeto Retenido", contribuyente, "Retenciones recibidas")
    add_contribuyente_alerts(alerts, ventas_df, "RUC Emisor", "Razon Social Emisor", contribuyente, "Ventas")
    add_contribuyente_alerts(alerts, ret_emitidas_df, "RUC Emisor", "Razon Social Emisor", contribuyente, "Retenciones emitidas")

    if ventas_df is not None and not ventas_df.empty and ret_emitidas_df is not None and not ret_emitidas_df.empty:
        venta_keys = {contribuyente_key(value) for value in ventas_df.get("RUC Emisor", []) if contribuyente_key(value)}
        ret_keys = {contribuyente_key(value) for value in ret_emitidas_df.get("RUC Emisor", []) if contribuyente_key(value)}
        if venta_keys and ret_keys and venta_keys != ret_keys:
            alerts.append({
                "Tipo": "NO COINCIDE EMISOR",
                "Detalle": "No coincide el emisor de la factura de venta con el emisor de las retenciones emitidas; el cuadro no concuerda.",
                "Fuente": "",
            })

    if not alerts:
        alerts.append({
            "Tipo": "OK",
            "Detalle": f"Cuadre correcto para {contribuyente.get('razon')} / {contribuyente.get('id')}.",
            "Fuente": "",
        })

    unique_alerts = []
    seen = set()
    for alert in alerts:
        key = tuple(clean_text(alert.get(column)) for column in ALERTAS_COLUMNS)
        if key in seen:
            continue
        seen.add(key)
        unique_alerts.append(alert)

    return pd.DataFrame(unique_alerts, columns=ALERTAS_COLUMNS)


def safe_filename(value):
    text = clean_text(value)
    text = re.sub(r'[<>:"/\\|?*]+', " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:120] or "Consolidado XML"


def excel_output_path(contribuyente):
    if EXCEL_FILE_MANUAL:
        return EXCEL_FILE
    razon = safe_filename(contribuyente.get("razon") or "Consolidado XML")
    identificacion = safe_filename(contribuyente.get("id") or "")
    name = f"{razon} {identificacion}".strip()
    return SALIDA_EXCEL_DIR / f"{name}.xlsx"


def infer_ats_context(compras_df=None, nc_df=None, ventas_df=None):
    razon = (
        first_df_value(compras_df, "Razon Social Comprador")
        or first_df_value(nc_df, "Razon Social Comprador")
        or first_df_value(ventas_df, "Razon Social Emisor")
        or "INFORMANTE"
    )
    identificacion = (
        first_df_value(compras_df, "Identificacion Comprador")
        or first_df_value(nc_df, "Identificacion Comprador")
        or first_df_value(ventas_df, "RUC Emisor")
    )

    date_values = []
    for df, column in (
        (compras_df, "Fecha Emision"),
        (nc_df, "Fecha Emision"),
        (ventas_df, "Fecha Emision"),
    ):
        if df is not None and not df.empty and column in df.columns:
            date_values.extend(df[column].tolist())
    inferred_year, inferred_month = most_common_period(date_values)

    year = configured_or_auto(ATS_ANIO, inferred_year or str(datetime.now().year))
    month = configured_or_auto(ATS_MES, inferred_month or f"{datetime.now().month:02d}")
    month = clean_fixed_digits(month, 2)

    establecimientos = set()
    if ventas_df is not None and not ventas_df.empty and "Establecimiento" in ventas_df.columns:
        for value in ventas_df["Establecimiento"]:
            establecimiento = clean_fixed_digits(value, 3)
            if establecimiento and establecimiento != "000":
                establecimientos.add(establecimiento)
    num_estab_inferido = f"{len(establecimientos):03d}" if establecimientos else ""

    return {
        "TipoIDInformante": clean_text(ATS_TIPO_ID_INFORMANTE) or "R",
        "IdInformante": clean_digits(configured_or_auto(ATS_ID_INFORMANTE, identificacion)),
        "razonSocial": configured_or_auto(ATS_RAZON_SOCIAL, razon),
        "Anio": year,
        "Mes": month,
        "regimenMicroempresa": ATS_REGIMEN_MICROEMPRESA,
        "numEstabRuc": configured_or_auto(ATS_NUM_ESTAB_RUC, num_estab_inferido),
        "totalVentas": xml_decimal(ATS_TOTAL_VENTAS),
        "codigoOperativo": ATS_CODIGO_OPERATIVO,
    }


def ats_header_values(context=None):
    if context:
        return dict(context)
    return {
        "TipoIDInformante": ATS_TIPO_ID_INFORMANTE,
        "IdInformante": ATS_ID_INFORMANTE,
        "razonSocial": ATS_RAZON_SOCIAL,
        "Anio": ATS_ANIO,
        "Mes": ATS_MES,
        "regimenMicroempresa": ATS_REGIMEN_MICROEMPRESA,
        "numEstabRuc": ATS_NUM_ESTAB_RUC,
        "totalVentas": xml_decimal(ATS_TOTAL_VENTAS),
        "codigoOperativo": ATS_CODIGO_OPERATIVO,
    }


def ats_xml_file_path(context=None):
    if ATS_XML_FILE_MANUAL:
        return Path(ATS_XML_FILE_MANUAL)
    context = context or ats_header_values()
    year = clean_integer_text(context.get("Anio"), strip_leading=False)
    month = clean_fixed_digits(context.get("Mes"), 2)
    if not year or not month:
        year = str(datetime.now().year)
        month = f"{datetime.now().month:02d}"
    return ATS_DIR / f"AT-{month}{year}.xml"


def unwrap_xml(file_path):
    tree = ET.parse(file_path)
    root = strip_namespaces(tree.getroot())

    numero_aut = get_text(root, ".//numeroAutorizacion")
    fecha_aut = safe_date_iso(get_text(root, ".//fechaAutorizacion"))

    comprobante = root.find(".//comprobante")
    if comprobante is not None:
        inner = "".join(comprobante.itertext()).strip()
        inner = re.sub(r"<!\[CDATA\[|\]\]>", "", inner).strip()
        if inner:
            inner_root = strip_namespaces(ET.fromstring(inner))
            return inner_root, numero_aut, fecha_aut

    return root, numero_aut, fecha_aut


def iter_impuestos_iva(factura_root):
    imps = factura_root.findall(".//infoFactura/totalConImpuestos/totalImpuesto")
    if imps:
        return imps
    imps = factura_root.findall(".//totalConImpuestos/totalImpuesto")
    if imps:
        return imps
    return factura_root.findall(".//detalles/detalle/impuestos/impuesto")


def parse_compra_row(factura, numero_aut, fecha_aut):
    info_trib = factura.find("infoTributaria")
    info_fac = factura.find("infoFactura")

    estab = get_text(info_trib, "estab")
    pto_emi = get_text(info_trib, "ptoEmi")
    secuencial = get_text(info_trib, "secuencial")
    numero_factura = f"{estab}-{pto_emi}-{secuencial}" if estab and pto_emi and secuencial else ""

    clave_acceso = get_text(info_trib, "claveAcceso")
    if not numero_aut:
        numero_aut = clave_acceso

    bases = {0.0: 0.0, 5.0: 0.0, 8.0: 0.0, 12.0: 0.0, 14.0: 0.0, 15.0: 0.0}
    ivas = {0.0: 0.0, 5.0: 0.0, 8.0: 0.0, 12.0: 0.0, 14.0: 0.0, 15.0: 0.0}
    no_objeto_iva = 0.0
    exento_iva = 0.0

    for imp in iter_impuestos_iva(factura):
        if get_text(imp, "codigo") != "2":
            continue
        codporc = get_text(imp, "codigoPorcentaje")
        base = to_float(get_text(imp, "baseImponible", "0"))
        valor = to_float(get_text(imp, "valor", "0"))

        if codporc in CODPORC_NO_OBJETO:
            no_objeto_iva += base
            continue
        if codporc in CODPORC_EXENTO:
            exento_iva += base
            continue

        rate = IVA_CODPORC_TO_RATE.get(codporc)
        if rate in bases:
            bases[rate] += base
            ivas[rate] += valor

    productos = [
        get_text(det, "descripcion")
        for det in factura.findall(".//detalles/detalle")
        if get_text(det, "descripcion")
    ]

    return {
        "Razon Social Comprador": get_text(info_fac, "razonSocialComprador"),
        "Identificacion Comprador": get_text(info_fac, "identificacionComprador"),
        "Razon Social Emisor": get_text(info_trib, "razonSocial"),
        "RUC Emisor": get_text(info_trib, "ruc"),
        "Fecha Emision": safe_date_iso(get_text(info_fac, "fechaEmision")),
        "Fecha Autorizacion": fecha_aut,
        "Numero de Factura": numero_factura,
        "Subtotal IVA 0%": bases[0.0],
        "IVA 0%": ivas[0.0],
        "Subtotal IVA 5%": bases[5.0],
        "IVA 5%": ivas[5.0],
        "Subtotal IVA 8%": bases[8.0],
        "IVA 8%": ivas[8.0],
        "Subtotal IVA 12%": bases[12.0],
        "IVA 12%": ivas[12.0],
        "Subtotal IVA 14%": bases[14.0],
        "IVA 14%": ivas[14.0],
        "Subtotal IVA 15%": bases[15.0],
        "IVA 15%": ivas[15.0],
        "No Objeto IVA": no_objeto_iva,
        "Exento IVA": exento_iva,
        "Total Sin Impuestos": to_float(get_text(info_fac, "totalSinImpuestos", "0")),
        "Total": to_float(get_text(info_fac, "importeTotal", "0")),
        "Producto 1": productos[0] if len(productos) > 0 else "",
        "Producto 2": productos[1] if len(productos) > 1 else "",
        "Producto 3": productos[2] if len(productos) > 2 else "",
        "Numero Autorizacion": numero_aut,
    }


def parse_nc_row(nc_root, numero_aut):
    info_trib = nc_root.find("infoTributaria")
    info_nc = nc_root.find("infoNotaCredito")

    estab = get_text(info_trib, "estab")
    pto_emi = get_text(info_trib, "ptoEmi")
    secuencial = get_text(info_trib, "secuencial")
    numero_nc = f"{estab}-{pto_emi}-{secuencial}" if estab and pto_emi and secuencial else ""

    clave_acceso = get_text(info_trib, "claveAcceso")
    if not numero_aut:
        numero_aut = clave_acceso

    bases = {0.0: 0.0, 5.0: 0.0, 8.0: 0.0, 12.0: 0.0, 14.0: 0.0, 15.0: 0.0}
    ivas = {0.0: 0.0, 5.0: 0.0, 8.0: 0.0, 12.0: 0.0, 14.0: 0.0, 15.0: 0.0}
    subtotal_no_objeto = 0.0
    no_objeto = 0.0
    subtotal_exento = 0.0

    for imp in nc_root.findall(".//totalConImpuestos/totalImpuesto"):
        if get_text(imp, "codigo") != "2":
            continue
        codporc = get_text(imp, "codigoPorcentaje")
        base = to_float(get_text(imp, "baseImponible", "0"))
        valor = to_float(get_text(imp, "valor", "0"))
        tarifa = get_text(imp, "tarifa")

        if codporc in CODPORC_NO_OBJETO:
            subtotal_no_objeto += base
            no_objeto += valor
            continue
        if codporc in CODPORC_EXENTO:
            subtotal_exento += base
            continue

        rate = to_float(tarifa, None) if tarifa else IVA_CODPORC_TO_RATE.get(codporc)
        if rate in bases:
            bases[rate] += base
            ivas[rate] += valor

    detalles = [
        get_text(det, "descripcion")
        for det in nc_root.findall(".//detalles/detalle")
        if get_text(det, "descripcion")
    ]

    return {
        "Razon Social Comprador": get_text(info_nc, "razonSocialComprador"),
        "Identificacion Comprador": get_text(info_nc, "identificacionComprador"),
        "Razon Social Emisor": get_text(info_trib, "razonSocial"),
        "RUC Emisor": get_text(info_trib, "ruc"),
        "Fecha Emision": safe_date_iso(get_text(info_nc, "fechaEmision")),
        "Numero de Nota de Crédito": numero_nc,
        "Numero Factura Modificada": get_text(info_nc, "numDocModificado"),
        "Subtotal IVA 0%": bases[0.0],
        "IVA 0%": ivas[0.0],
        "Subtotal IVA 5%": bases[5.0],
        "IVA 5%": ivas[5.0],
        "Subtotal IVA 8%": bases[8.0],
        "IVA 8%": ivas[8.0],
        "Subtotal IVA 12%": bases[12.0],
        "IVA 12%": ivas[12.0],
        "Subtotal IVA 14%": bases[14.0],
        "IVA 14%": ivas[14.0],
        "Subtotal IVA 15%": bases[15.0],
        "IVA 15%": ivas[15.0],
        "Subtotal No Objeto": subtotal_no_objeto,
        "No Objeto": no_objeto,
        "Subtotal Exento": subtotal_exento,
        "Total": to_float(get_text(info_nc, "valorModificacion", "0")),
        "Numero Autorización": numero_aut,
        "Detalle 1": detalles[0] if len(detalles) > 0 else "",
        "Detalle 2": detalles[1] if len(detalles) > 1 else "",
        "Detalle 3": detalles[2] if len(detalles) > 2 else "",
    }


def tipo_retencion_label(codigo):
    return TIPO_RETENCION_MAP.get((codigo or "").strip(), codigo or "")


def build_ret_doc_uid(ruc, estab, pto_emi, secuencial, fecha_emision, clave, numero_aut):
    if numero_aut:
        return f"AUT:{numero_aut}"
    if clave:
        return f"CLA:{clave}"
    return f"ALT:{ruc}|{estab}|{pto_emi}|{secuencial}|{fecha_emision}"


def parse_retencion_rows(ret_root, numero_aut):
    info_trib = ret_root.find("infoTributaria")
    info_comp = ret_root.find("infoCompRetencion")

    estab = get_text(info_trib, "estab")
    pto_emi = get_text(info_trib, "ptoEmi")
    secuencial = get_text(info_trib, "secuencial")
    clave_acceso = get_text(info_trib, "claveAcceso")
    fecha_emision = safe_date_iso(get_text(info_comp, "fechaEmision"))

    if not numero_aut:
        numero_aut = clave_acceso

    doc_uid = build_ret_doc_uid(
        get_text(info_trib, "ruc"), estab, pto_emi, secuencial, fecha_emision, clave_acceso, numero_aut
    )

    base_header = {
        "Razon Social Emisor": get_text(info_trib, "razonSocial"),
        "RUC Emisor": get_text(info_trib, "ruc"),
        "Establecimiento": estab,
        "Punto Emision": pto_emi,
        "Numero Comprobante Retencion": secuencial,
        "Razon Social Sujeto Retenido": get_text(info_comp, "razonSocialSujetoRetenido"),
        "Identificacion Sujeto Retenido": get_text(info_comp, "identificacionSujetoRetenido"),
        "Fecha Emision": fecha_emision,
        "Periodo Fiscal": get_text(info_comp, "periodoFiscal"),
        "Clave Acceso": clave_acceso,
        "Numero Autorizacion": numero_aut,
    }

    rows = []
    docs = ret_root.findall(".//docsSustento/docSustento")
    if docs:
        for doc in docs:
            doc_data = {
                "Numero Doc Sustento": get_text(doc, "numDocSustento"),
                "Fecha Doc Sustento": safe_date_iso(get_text(doc, "fechaEmisionDocSustento")),
                "Total sin Impuestos": to_float(get_text(doc, "totalSinImpuestos", "0")),
                "Importe Total": to_float(get_text(doc, "importeTotal", "0")),
            }
            rets = doc.findall(".//retenciones/retencion")
            if not rets:
                rows.append({
                    **base_header,
                    **doc_data,
                    "Tipo Retencion": "",
                    "Codigo Retencion": "",
                    "Base Imponible": 0.0,
                    "Porcentaje Retencion": 0.0,
                    "Valor Retenido": 0.0,
                })
                continue
            for ret in rets:
                rows.append({
                    **base_header,
                    **doc_data,
                    "Tipo Retencion": tipo_retencion_label(get_text(ret, "codigo")),
                    "Codigo Retencion": get_text(ret, "codigoRetencion"),
                    "Base Imponible": to_float(get_text(ret, "baseImponible", "0")),
                    "Porcentaje Retencion": to_float(get_text(ret, "porcentajeRetener", "0")),
                    "Valor Retenido": to_float(get_text(ret, "valorRetenido", "0")),
                })
        return rows, doc_uid

    impuestos = ret_root.findall(".//impuestos/impuesto")
    if impuestos:
        for imp in impuestos:
            rows.append({
                **base_header,
                "Numero Doc Sustento": get_text(imp, "numDocSustento"),
                "Fecha Doc Sustento": safe_date_iso(get_text(imp, "fechaEmisionDocSustento")),
                "Total sin Impuestos": "",
                "Importe Total": "",
                "Tipo Retencion": tipo_retencion_label(get_text(imp, "codigo")),
                "Codigo Retencion": get_text(imp, "codigoRetencion"),
                "Base Imponible": to_float(get_text(imp, "baseImponible", "0")),
                "Porcentaje Retencion": to_float(get_text(imp, "porcentajeRetener", "0")),
                "Valor Retenido": to_float(get_text(imp, "valorRetenido", "0")),
            })
        return rows, doc_uid

    rows.append({
        **base_header,
        "Numero Doc Sustento": "",
        "Fecha Doc Sustento": "",
        "Total sin Impuestos": "",
        "Importe Total": "",
        "Tipo Retencion": "",
        "Codigo Retencion": "",
        "Base Imponible": 0.0,
        "Porcentaje Retencion": 0.0,
        "Valor Retenido": 0.0,
    })
    return rows, doc_uid


def pdf_money_to_float(value):
    text = clean_text(value)
    if not text:
        return 0.0
    text = re.sub(r"[^\d,.-]", "", text)
    if "," in text and "." in text:
        text = text.replace(",", "")
    return to_float(text)


def regex_first(pattern, text, default="", flags=re.IGNORECASE | re.DOTALL):
    match = re.search(pattern, text, flags)
    return clean_text(match.group(1)) if match else default


def pdf_amount_by_label(text, label):
    if normalize_header(label).startswith("iva "):
        line_pattern = rf"^\s*{re.escape(label)}\s+([0-9]+(?:[.,][0-9]{{2}})?)"
        value = regex_first(line_pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if value:
            return pdf_money_to_float(value)
    pattern = rf"{re.escape(label)}\s+([0-9]+(?:[.,][0-9]{{2}})?)"
    return pdf_money_to_float(regex_first(pattern, text))


def clean_pdf_cell(value):
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def extract_pdf_products_and_payments(pdf_path):
    products = []
    payment_code = ""

    import pdfplumber

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for row in table[1:]:
                    cells = [clean_pdf_cell(cell) for cell in row]
                    if len(cells) >= 4 and re.match(r"^\d+(?:[.,]\d+)?$", cells[2] or ""):
                        description = cells[3]
                        if description and "descripcion" not in normalize_header(description):
                            products.append(description)
                    if len(cells) >= 2 and re.match(r"^\d{2}\s*-", cells[0] or ""):
                        payment_code = clean_digits(cells[0][:2])

    return products, payment_code


def extract_pdf_text(pdf_path):
    import pdfplumber

    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text(x_tolerance=1, y_tolerance=3) or "")
    return "\n".join(parts)


def pdf_issuer_name(text):
    lines = [clean_pdf_cell(line) for line in text.splitlines() if clean_pdf_cell(line)]
    for idx, line in enumerate(lines):
        if normalize_header(line).startswith("ambiente"):
            for candidate in lines[idx + 1:idx + 4]:
                normalized = normalize_header(candidate)
                if "emision" not in normalized or "direccion" in normalized:
                    continue
                value = re.split(r"\s+EMISI[ÓO]N", candidate, maxsplit=1, flags=re.IGNORECASE)[0]
                if value:
                    return clean_pdf_cell(value)

    for line in lines:
        normalized = normalize_header(line)
        if "emision" not in normalized or "direccion" in normalized or normalized.startswith("fecha"):
            continue
        value = re.split(r"\s+EMISI[ÓO]N", line, maxsplit=1, flags=re.IGNORECASE)[0]
        if value:
            return clean_pdf_cell(value)

    return clean_pdf_cell(regex_first(r"^\s*(.+?)\s+FECHA Y HORA DE", text, flags=re.IGNORECASE | re.MULTILINE))


def pdf_client_name(text):
    after_label = regex_first(
        r"Raz[oó]n Social / Nombres y Apellidos:\s*(.*?)\s*Identificaci[oó]n",
        text,
    )
    before_label = regex_first(
        r"\n([A-Z0-9][^\n]+)\nRaz[oó]n Social / Nombres y Apellidos:",
        text,
    )
    before_digits = clean_digits(before_label)
    before_is_authorization = len(before_digits) >= 30 and before_digits == re.sub(r"\D", "", clean_text(before_label))
    if before_label and len(before_label) > len(after_label) and not before_is_authorization:
        return clean_pdf_cell(f"{before_label} {after_label}")
    return clean_pdf_cell(after_label)


def validate_venta_row(row, context=None):
    required = [
        "RUC Emisor",
        "Identificacion Cliente",
        "Fecha Emision",
        "Numero de Factura",
        "Numero Autorizacion",
        "tpIdCliente",
        "idCliente",
        "tipoComprobante",
        "tipoEmision",
    ]
    missing = [field for field in required if not clean_text(row.get(field))]
    if missing:
        row["Exportar XML"] = "NO"
        row["Estado ATS"] = "REVISAR"
        row["Observacion"] = "Faltan campos obligatorios: " + ", ".join(missing)
        return row

    if FILTRAR_ATS_POR_PERIODO and context:
        parsed = parse_date_any(clean_text(row.get("Fecha Emision")))
        if parsed:
            expected_year = int(clean_text(context.get("Anio")))
            expected_month = int(clean_text(context.get("Mes")))
            if parsed.year != expected_year or parsed.month != expected_month:
                row["Exportar XML"] = "NO"
                row["Estado ATS"] = "FUERA_PERIODO"
                row["Observacion"] = f"Documento fuera del periodo ATS configurado {expected_month:02d}/{expected_year}."
    return row


def parse_venta_pdf_row(pdf_path):
    text = extract_pdf_text(pdf_path)
    products, payment_code = extract_pdf_products_and_payments(pdf_path)

    numero_factura = regex_first(r"No\.\s*([0-9]{3}-[0-9]{3}-[0-9]{9})", text)
    establecimiento, punto_emision, secuencial = split_document_number(numero_factura)
    ruc_emisor = regex_first(r"R\.?U\.?C\.?:\s*([0-9]{13})", text)
    identificacion_cliente = regex_first(r"Identificaci[oó]n:?\s*([0-9]{10,13})", text)
    fecha_emision = regex_first(r"\bFecha(?:\s+emisi[oó]n)?:?\s*([0-9]{2}/[0-9]{2}/[0-9]{4})", text)
    fecha_aut = regex_first(r"FECHA Y HORA DE AUTORIZACI[ÓO]N:?\s*([0-9]{2}/[0-9]{2}/[0-9]{4})", text)
    if not fecha_aut:
        fecha_aut = regex_first(r"([0-9]{2}/[0-9]{2}/[0-9]{4}\s+[0-9]{2}:[0-9]{2}:[0-9]{2})\s*AUTORIZACI", text)
    numero_aut = regex_first(r"N[ÚU]MERO DE AUTORIZACI[ÓO]N\s+([0-9]{37,49})", text)

    subtotal_0 = pdf_amount_by_label(text, "SUBTOTAL IVA 0%") or pdf_amount_by_label(text, "SUBTOTAL 0%")
    subtotal_5 = pdf_amount_by_label(text, "SUBTOTAL IVA 5%") or pdf_amount_by_label(text, "SUBTOTAL 5%")
    subtotal_8 = pdf_amount_by_label(text, "SUBTOTAL IVA 8%") or pdf_amount_by_label(text, "SUBTOTAL 8%")
    subtotal_12 = pdf_amount_by_label(text, "SUBTOTAL IVA 12%") or pdf_amount_by_label(text, "SUBTOTAL 12%")
    subtotal_14 = pdf_amount_by_label(text, "SUBTOTAL IVA 14%") or pdf_amount_by_label(text, "SUBTOTAL 14%")
    subtotal_15 = pdf_amount_by_label(text, "SUBTOTAL IVA 15%") or pdf_amount_by_label(text, "SUBTOTAL 15%")
    iva_5 = pdf_amount_by_label(text, "IVA 5%")
    iva_8 = pdf_amount_by_label(text, "IVA 8%")
    iva_12 = pdf_amount_by_label(text, "IVA 12%")
    iva_14 = pdf_amount_by_label(text, "IVA 14%")
    iva_15 = pdf_amount_by_label(text, "IVA 15%")
    no_objeto = pdf_amount_by_label(text, "SUBTOTAL NO OBJETO DE IVA")
    exento = pdf_amount_by_label(text, "SUBTOTAL EXENTO DE IVA")
    total_sin_impuestos = pdf_amount_by_label(text, "SUBTOTAL SIN IMPUESTOS")
    ice = pdf_amount_by_label(text, "ICE")
    total = pdf_amount_by_label(text, "VALOR TOTAL")
    if not payment_code:
        payment_code = regex_first(r"Forma de pago\s+Valor\s+([0-9]{2})\s*-", text)

    row = {column: "" for column in VENTAS_COLUMNS}
    row.update({
        "Fuente PDF": Path(pdf_path).name,
        "Exportar XML": "SI",
        "Estado ATS": "OK",
        "Observacion": "",
        "Razon Social Emisor": pdf_issuer_name(text),
        "RUC Emisor": ruc_emisor,
        "Razon Social Cliente": pdf_client_name(text),
        "Identificacion Cliente": identificacion_cliente,
        "Fecha Emision": date_ddmmyyyy(fecha_emision),
        "Fecha Autorizacion": date_ddmmyyyy(fecha_aut),
        "Numero de Factura": numero_factura,
        "Establecimiento": establecimiento,
        "Punto Emision": punto_emision,
        "Secuencial": secuencial,
        "Numero Autorizacion": numero_aut,
        "Subtotal IVA 0%": subtotal_0,
        "IVA 0%": 0.0,
        "Subtotal IVA 5%": subtotal_5,
        "Subtotal IVA 8%": subtotal_8,
        "Subtotal IVA 12%": subtotal_12,
        "Subtotal IVA 14%": subtotal_14,
        "Subtotal IVA 15%": subtotal_15,
        "IVA 5%": iva_5,
        "IVA 8%": iva_8,
        "IVA 12%": iva_12,
        "IVA 14%": iva_14,
        "IVA 15%": iva_15,
        "No Objeto IVA": no_objeto,
        "Exento IVA": exento,
        "ICE": ice,
        "Total Sin Impuestos": total_sin_impuestos,
        "Total": total,
        "Forma Pago": payment_code,
        "Producto 1": products[0] if len(products) > 0 else "",
        "Producto 2": products[1] if len(products) > 1 else "",
        "Producto 3": products[2] if len(products) > 2 else "",
        "tpIdCliente": tipo_id_cliente_from_identificacion(identificacion_cliente),
        "idCliente": clean_digits(identificacion_cliente),
        "parteRelVtas": "NO",
        "tipoComprobante": "18",
        "tipoEmision": ATS_TIPO_EMISION_VENTAS_PDF,
        "numeroComprobantes": 1,
        "valorRetIva": 0.0,
        "valorRetRenta": 0.0,
    })
    return validate_venta_row(row)


def is_retencion_pdf_text(text):
    normalized = normalize_header(text)
    return "comprobante de retencion" in normalized


def is_factura_pdf_text(text):
    normalized = normalize_header(text)
    return "factura" in normalized and "comprobante de retencion" not in normalized


def pdf_retencion_sujeto_name(text):
    return clean_pdf_cell(regex_first(r"Raz[oó]n Social / Nombres y Apellidos:\s*(.*?)\s*Identificaci[oó]n", text, flags=re.IGNORECASE))


def pdf_retencion_issuer_name(text):
    lines = [clean_pdf_cell(line) for line in text.splitlines() if clean_pdf_cell(line)]
    for idx, line in enumerate(lines):
        if re.match(r"^[0-9]{2}/[0-9]{2}/[0-9]{4}\s+[0-9]{2}:[0-9]{2}:[0-9]{2}", line):
            for candidate in lines[idx + 1:idx + 4]:
                normalized = normalize_header(candidate)
                if "autorizacion" not in normalized:
                    continue
                if normalized.startswith("numero de") or normalized.startswith("fecha y hora"):
                    continue
                value = re.split(r"\s+AUTORIZACI", candidate, maxsplit=1, flags=re.IGNORECASE)[0]
                if value:
                    return clean_pdf_cell(value)

    for line in lines:
        normalized = normalize_header(line)
        if "autorizacion" not in normalized:
            continue
        if normalized.startswith("numero de") or normalized.startswith("fecha y hora"):
            continue
        value = re.split(r"\s+AUTORIZACI", line, maxsplit=1, flags=re.IGNORECASE)[0]
        if value:
            return clean_pdf_cell(value)

    return clean_pdf_cell(regex_first(r"^\s*(.+?)\s+FECHA Y HORA", text, flags=re.IGNORECASE | re.MULTILINE))


def parse_retencion_emitida_pdf_rows(pdf_path, text=None):
    text = text or extract_pdf_text(pdf_path)
    numero_retencion = regex_first(r"No\.\s*([0-9]{3}-[0-9]{3}-[0-9]{9})", text)
    ruc_emisor = regex_first(r"R\.?U\.?C\.?:\s*([0-9]{13})", text, flags=re.IGNORECASE)
    numero_aut = regex_first(r"N[ÚU]MERO DE AUTORIZACI[ÓO]N\s+([0-9]{37,49})", text, flags=re.IGNORECASE)
    fecha_aut = regex_first(r"([0-9]{2}/[0-9]{2}/[0-9]{4}\s+[0-9]{2}:[0-9]{2}:[0-9]{2})", text)
    sujeto_id = regex_first(r"Identificaci[oó]n\s+([0-9]{10,13})", text, flags=re.IGNORECASE)
    fecha_emision = regex_first(r"\bFecha\s+([0-9]{2}/[0-9]{2}/[0-9]{4})", text, flags=re.IGNORECASE)
    periodo_fiscal = regex_first(r"\b([0-9]{2}/[0-9]{4})\b", text)
    doc_sustento = regex_first(r"\b([0-9]{15})\b", text)
    if doc_sustento:
        doc_sustento = f"{doc_sustento[:3]}-{doc_sustento[3:6]}-{doc_sustento[6:]}"
    fecha_doc = regex_first(r"FACTURA\s+([0-9]{2}/[0-9]{2}/[0-9]{4})", text, flags=re.IGNORECASE)

    base_header = {
        "Fuente PDF": Path(pdf_path).name,
        "Estado": "OK",
        "Observacion": "",
        "Razon Social Emisor": pdf_retencion_issuer_name(text),
        "RUC Emisor": ruc_emisor,
        "Numero Comprobante Retencion": numero_retencion,
        "Razon Social Sujeto Retenido": pdf_retencion_sujeto_name(text),
        "Identificacion Sujeto Retenido": sujeto_id,
        "Fecha Emision": date_ddmmyyyy(fecha_emision),
        "Periodo Fiscal": periodo_fiscal,
        "Numero Doc Sustento": doc_sustento,
        "Fecha Doc Sustento": date_ddmmyyyy(fecha_doc),
        "Numero Autorizacion": numero_aut,
    }

    rows = []
    renta_match = re.search(r"([0-9]+(?:[.,][0-9]{2})?)\s+([0-9]+(?:[.,][0-9]{2})?)\s+([0-9]+(?:[.,][0-9]{2})?)\s*\n\s*Renta", text, flags=re.IGNORECASE)
    if not renta_match:
        renta_match = re.search(r"Renta.*?([0-9]+(?:[.,][0-9]{2})?)\s+([0-9]+(?:[.,][0-9]{2})?)\s+([0-9]+(?:[.,][0-9]{2})?)", text, flags=re.IGNORECASE | re.DOTALL)
    if renta_match:
        base, porcentaje, valor = renta_match.groups()
        rows.append({
            **base_header,
            "Tipo Retencion": "RENTA",
            "Codigo Retencion": "",
            "Base Imponible": pdf_money_to_float(base),
            "Porcentaje Retencion": pdf_money_to_float(porcentaje),
            "Valor Retenido": pdf_money_to_float(valor),
        })

    iva_matches = re.findall(r"([0-9]+(?:[.,][0-9]{2})?)\s+IVA\s+([0-9]+(?:[.,][0-9]{2})?)\s+([0-9]+(?:[.,][0-9]{2})?)", text, flags=re.IGNORECASE)
    for base, porcentaje, valor in iva_matches:
        rows.append({
            **base_header,
            "Tipo Retencion": "IVA",
            "Codigo Retencion": "",
            "Base Imponible": pdf_money_to_float(base),
            "Porcentaje Retencion": pdf_money_to_float(porcentaje),
            "Valor Retenido": pdf_money_to_float(valor),
        })

    if not rows:
        row = {column: "" for column in RET_EMITIDAS_COLUMNS}
        row.update(base_header)
        row.update({
            "Estado": "REVISAR",
            "Observacion": "No se detectaron lineas de retencion en el PDF.",
            "Base Imponible": 0.0,
            "Porcentaje Retencion": 0.0,
            "Valor Retenido": 0.0,
        })
        rows.append(row)

    return rows


def process_pdf_ventas_y_retenciones_emitidas():
    ventas_rows = []
    ret_emitidas_rows = []
    errors = []
    for pdf_file in list_pdf_files(PDF_VENTAS_FOLDER, PDF_RECURSIVE):
        try:
            text = extract_pdf_text(pdf_file)
            if is_retencion_pdf_text(text):
                ret_emitidas_rows.extend(parse_retencion_emitida_pdf_rows(pdf_file, text=text))
            elif is_factura_pdf_text(text):
                ventas_rows.append(parse_venta_pdf_row(pdf_file))
            else:
                errors.append(f"{pdf_file.name}: PDF no reconocido como factura de venta ni retencion emitida.")
        except Exception as exc:
            errors.append(f"{pdf_file.name}: {exc}")

    ventas_df = pd.DataFrame(ventas_rows)
    if ventas_df.empty:
        ventas_df = pd.DataFrame(columns=VENTAS_COLUMNS)
    else:
        ventas_df = ensure_columns(ventas_df, VENTAS_COLUMNS, VENTAS_NUMERIC_COLS)
        dedup_key = [column for column in ["RUC Emisor", "Numero de Factura", "Numero Autorizacion"] if column in ventas_df.columns]
        if dedup_key:
            ventas_df = ventas_df.drop_duplicates(subset=dedup_key, keep="first")
        ventas_df = apply_decimal_text(ventas_df, VENTAS_NUMERIC_COLS)

    ret_emitidas_df = pd.DataFrame(ret_emitidas_rows)
    if ret_emitidas_df.empty:
        ret_emitidas_df = pd.DataFrame(columns=RET_EMITIDAS_COLUMNS)
    else:
        ret_emitidas_df = ensure_columns(ret_emitidas_df, RET_EMITIDAS_COLUMNS, RET_EMITIDAS_NUMERIC_COLS)
        dedup_key = [column for column in ["RUC Emisor", "Numero Comprobante Retencion", "Tipo Retencion", "Valor Retenido"] if column in ret_emitidas_df.columns]
        if dedup_key:
            ret_emitidas_df = ret_emitidas_df.drop_duplicates(subset=dedup_key, keep="first")
        ret_emitidas_df = apply_decimal_text(ret_emitidas_df, RET_EMITIDAS_NUMERIC_COLS)

    return ventas_df, ret_emitidas_df, errors


def process_pdf_ventas():
    ventas_df, _, errors = process_pdf_ventas_y_retenciones_emitidas()
    return ventas_df, errors


def validate_ventas_dataframe(ventas_df, context=None):
    if ventas_df.empty:
        return ventas_df
    rows = []
    for _, row in ventas_df.iterrows():
        row_dict = row.to_dict()
        rows.append(validate_venta_row(row_dict, context))
    return pd.DataFrame(rows)[VENTAS_COLUMNS]


def ensure_columns(df, columns, numeric_cols):
    for column in columns:
        if column not in df.columns:
            df[column] = 0.0 if column in numeric_cols else ""
    return df[columns]


def apply_decimal_text(df, numeric_cols):
    if DECIMALES_CON_COMA:
        for column in numeric_cols:
            if column in df.columns:
                df[column] = df[column].apply(format_decimal_comma)
    return df


def prepare_dataframe(rows, columns, numeric_cols, dedup_key):
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=columns), 0

    df = ensure_columns(df, columns, numeric_cols)
    dedup_key = [column for column in dedup_key if column in df.columns]
    skipped = 0
    if dedup_key:
        before = len(df)
        df = df.drop_duplicates(subset=dedup_key, keep="first")
        skipped = before - len(df)
    df = apply_decimal_text(df, numeric_cols)
    return df, skipped


def empty_ats_row(source, exportar="SI", estado="OK", observacion="", context=None):
    row = {column: "" for column in ATS_COLUMNS}
    row.update(ats_header_values(context))
    row.update({
        "Exportar XML": exportar,
        "Estado ATS": estado,
        "Observacion": observacion,
        "Fuente": source,
        "codSustento": ATS_COD_SUSTENTO_DEFAULT,
        "parteRel": ATS_PARTE_REL_DEFAULT,
        "baseNoGraIva": "0.00",
        "baseImponible": "0.00",
        "baseImpGrav": "0.00",
        "baseImpExe": "0.00",
        "montoIce": "0.00",
        "montoIva": "0.00",
        "valRetBien10": "0.00",
        "valRetServ20": "0.00",
        "valorRetBienes": "0.00",
        "valRetServ50": "0.00",
        "valorRetServicios": "0.00",
        "valRetServ100": "0.00",
        "valorRetencionNc": "0.00",
        "totbasesImpReemb": "0.00",
        "pagoLocExt": ATS_PAGO_LOC_EXT_DEFAULT,
        "paisEfecPago": ATS_PAIS_EFEC_PAGO_DEFAULT,
        "aplicConvDobTrib": ATS_APLIC_CONV_DOB_TRIB_DEFAULT,
        "pagExtSujRetNorLeg": ATS_PAG_EXT_SUJ_RET_NOR_LEG_DEFAULT,
        "baseImpAir": "",
        "porcentajeAir": "",
        "valRetAir": "",
    })
    return row


def retencion_iva_field(percent):
    rounded = int(round(to_float(percent)))
    return {
        10: "valRetBien10",
        20: "valRetServ20",
        30: "valorRetBienes",
        50: "valRetServ50",
        70: "valorRetServicios",
        100: "valRetServ100",
    }.get(rounded)


def air_rows_to_text(rows):
    parts = []
    for row in rows:
        values = [
            clean_text(row.get("codRetAir")),
            xml_decimal(row.get("baseImpAir")),
            xml_decimal(row.get("porcentajeAir")),
            xml_decimal(row.get("valRetAir")),
        ]
        if values[0]:
            parts.append("|".join(values))
    return ";".join(parts)


def air_row_from_ats_columns(row):
    if not clean_text(row.get("codRetAir")):
        return None
    return {
        "codRetAir": clean_text(row.get("codRetAir")),
        "baseImpAir": row.get("baseImpAir"),
        "porcentajeAir": row.get("porcentajeAir"),
        "valRetAir": row.get("valRetAir"),
    }


def parse_air_text(value):
    rows = []
    text = clean_text(value)
    if not text:
        return rows
    for item in text.split(";"):
        fields = [part.strip() for part in item.split("|")]
        if len(fields) < 4 or not fields[0]:
            continue
        rows.append({
            "codRetAir": fields[0],
            "baseImpAir": fields[1],
            "porcentajeAir": fields[2],
            "valRetAir": fields[3],
        })
    return rows


def build_retenciones_index(ret_df):
    index = {}
    if ret_df.empty:
        return index

    for _, ret in ret_df.iterrows():
        key = document_match_key(ret.get("Numero Doc Sustento"))
        if not key:
            continue

        item = index.setdefault(key, {
            "iva": {
                "valRetBien10": 0.0,
                "valRetServ20": 0.0,
                "valorRetBienes": 0.0,
                "valRetServ50": 0.0,
                "valorRetServicios": 0.0,
                "valRetServ100": 0.0,
            },
            "air": [],
            "ret_doc": ret,
            "used": False,
        })

        tipo = clean_text(ret.get("Tipo Retencion")).upper()
        valor = to_float(ret.get("Valor Retenido"))

        if tipo == "IVA":
            field = retencion_iva_field(ret.get("Porcentaje Retencion"))
            if field:
                item["iva"][field] += valor
        elif tipo == "RENTA":
            item["air"].append({
                "codRetAir": clean_text(ret.get("Codigo Retencion")),
                "baseImpAir": ret.get("Base Imponible"),
                "porcentajeAir": ret.get("Porcentaje Retencion"),
                "valRetAir": ret.get("Valor Retenido"),
            })

    return index


def apply_retencion_to_ats_row(row, ret_info):
    if not ret_info:
        return row

    for field, value in ret_info["iva"].items():
        row[field] = xml_decimal(value)

    air_rows = ret_info["air"]
    if air_rows:
        first_air = air_rows[0]
        row["codRetAir"] = clean_text(first_air.get("codRetAir"))
        row["baseImpAir"] = xml_decimal(first_air.get("baseImpAir"))
        row["porcentajeAir"] = xml_decimal(first_air.get("porcentajeAir"))
        row["valRetAir"] = xml_decimal(first_air.get("valRetAir"))
        row["air_detalle"] = air_rows_to_text(air_rows[1:])

    ret_doc = ret_info["ret_doc"]
    row["estabRetencion1"] = clean_fixed_digits(ret_doc.get("Establecimiento"), 3)
    row["ptoEmiRetencion1"] = clean_fixed_digits(ret_doc.get("Punto Emision"), 3)
    row["secRetencion1"] = clean_integer_text(ret_doc.get("Numero Comprobante Retencion"))
    row["autRetencion1"] = clean_digits(ret_doc.get("Numero Autorizacion"))
    row["fechaEmiRet1"] = date_ddmmyyyy(ret_doc.get("Fecha Emision"))
    ret_info["used"] = True
    return row


def compra_tiene_iva(row):
    return (
        sum_columns(row, [
            "Subtotal IVA 5%",
            "Subtotal IVA 8%",
            "Subtotal IVA 12%",
            "Subtotal IVA 14%",
            "Subtotal IVA 15%",
            "IVA 5%",
            "IVA 8%",
            "IVA 12%",
            "IVA 14%",
            "IVA 15%",
        ]) > 0
    )


def cod_sustento_desde_valores(row):
    return "01" if compra_tiene_iva(row) else "02"


def ats_row_from_compra(compra, retenciones_index, context=None):
    numero_factura = compra.get("Numero de Factura")
    establecimiento, punto_emision, secuencial = split_document_number(numero_factura)
    key = document_match_key(numero_factura)

    row = empty_ats_row("Compra", context=context)
    row.update({
        "codSustento": cod_sustento_desde_valores(compra),
        "tpIdProv": tipo_id_from_identificacion(compra.get("RUC Emisor")),
        "idProv": clean_digits(compra.get("RUC Emisor")),
        "tipoComprobante": "01",
        "fechaRegistro": date_ddmmyyyy(compra.get("Fecha Emision")),
        "establecimiento": establecimiento,
        "puntoEmision": punto_emision,
        "secuencial": secuencial,
        "fechaEmision": date_ddmmyyyy(compra.get("Fecha Emision")),
        "autorizacion": clean_digits(compra.get("Numero Autorizacion")),
        "baseNoGraIva": xml_decimal(compra.get("No Objeto IVA")),
        "baseImponible": xml_decimal(compra.get("Subtotal IVA 0%")),
        "baseImpGrav": xml_decimal(sum_columns(compra, [
            "Subtotal IVA 5%",
            "Subtotal IVA 8%",
            "Subtotal IVA 12%",
            "Subtotal IVA 14%",
            "Subtotal IVA 15%",
        ])),
        "baseImpExe": xml_decimal(compra.get("Exento IVA")),
        "montoIce": "0.00",
        "montoIva": xml_decimal(sum_columns(compra, [
            "IVA 5%",
            "IVA 8%",
            "IVA 12%",
            "IVA 14%",
            "IVA 15%",
        ])),
    })
    apply_retencion_to_ats_row(row, retenciones_index.get(key))
    return validate_ats_row(row, context)


def ats_row_from_nota_credito(nota, retenciones_index, context=None):
    numero_nc = first_existing(nota, ["Numero de Nota de Credito", "Numero de Nota de CrÃ©dito"])
    establecimiento, punto_emision, secuencial = split_document_number(numero_nc)
    key = document_match_key(numero_nc)
    estab_mod, pto_mod, sec_mod = split_document_number(nota.get("Numero Factura Modificada"))

    row = empty_ats_row("NotaCredito", context=context)
    row.update({
        "codSustento": cod_sustento_desde_valores(nota),
        "tpIdProv": tipo_id_from_identificacion(nota.get("RUC Emisor")),
        "idProv": clean_digits(nota.get("RUC Emisor")),
        "tipoComprobante": "04",
        "fechaRegistro": date_ddmmyyyy(nota.get("Fecha Emision")),
        "establecimiento": establecimiento,
        "puntoEmision": punto_emision,
        "secuencial": secuencial,
        "fechaEmision": date_ddmmyyyy(nota.get("Fecha Emision")),
        "autorizacion": clean_digits(first_existing(nota, ["Numero Autorizacion", "Numero AutorizaciÃ³n"])),
        "baseNoGraIva": xml_decimal(first_existing(nota, ["Subtotal No Objeto", "No Objeto"])),
        "baseImponible": xml_decimal(nota.get("Subtotal IVA 0%")),
        "baseImpGrav": xml_decimal(sum_columns(nota, [
            "Subtotal IVA 5%",
            "Subtotal IVA 8%",
            "Subtotal IVA 12%",
            "Subtotal IVA 14%",
            "Subtotal IVA 15%",
        ])),
        "baseImpExe": xml_decimal(nota.get("Subtotal Exento")),
        "montoIce": "0.00",
        "montoIva": xml_decimal(sum_columns(nota, [
            "IVA 5%",
            "IVA 8%",
            "IVA 12%",
            "IVA 14%",
            "IVA 15%",
        ])),
        "valorRetencionNc": "0.00",
        "docModificado": "01" if estab_mod and pto_mod and sec_mod else "",
        "estabModificado": estab_mod,
        "ptoEmiModificado": pto_mod,
        "secModificado": sec_mod,
    })
    apply_retencion_to_ats_row(row, retenciones_index.get(key))
    return validate_ats_row(row, context)


def ats_row_from_retencion_pendiente(ret_info, context=None):
    if isinstance(ret_info, dict) and "ret_doc" in ret_info:
        ret = ret_info["ret_doc"]
        air_rows = ret_info.get("air", [])
        iva_values = ret_info.get("iva", {})
    else:
        ret = ret_info
        air_rows = []
        iva_values = {}

    establecimiento, punto_emision, secuencial = split_document_number(ret.get("Numero Doc Sustento"))
    row = empty_ats_row(
        "RetencionSinCompra",
        exportar="NO",
        estado="REVISAR",
        observacion="Retencion sin compra/nota relacionada en el consolidado; complete bases antes de exportar.",
        context=context,
    )
    row.update({
        "tpIdProv": tipo_id_from_identificacion(ret.get("Identificacion Sujeto Retenido")),
        "idProv": clean_digits(ret.get("Identificacion Sujeto Retenido")),
        "tipoComprobante": "01",
        "fechaRegistro": date_ddmmyyyy(ret.get("Fecha Doc Sustento") or ret.get("Fecha Emision")),
        "establecimiento": establecimiento,
        "puntoEmision": punto_emision,
        "secuencial": secuencial,
        "fechaEmision": date_ddmmyyyy(ret.get("Fecha Doc Sustento") or ret.get("Fecha Emision")),
        "estabRetencion1": clean_fixed_digits(ret.get("Establecimiento"), 3),
        "ptoEmiRetencion1": clean_fixed_digits(ret.get("Punto Emision"), 3),
        "secRetencion1": clean_integer_text(ret.get("Numero Comprobante Retencion")),
        "autRetencion1": clean_digits(ret.get("Numero Autorizacion")),
        "fechaEmiRet1": date_ddmmyyyy(ret.get("Fecha Emision")),
    })

    for field, value in iva_values.items():
        row[field] = xml_decimal(value)

    if air_rows:
        first_air = air_rows[0]
        row.update({
            "codRetAir": clean_text(first_air.get("codRetAir")),
            "baseImpAir": xml_decimal(first_air.get("baseImpAir")),
            "porcentajeAir": xml_decimal(first_air.get("porcentajeAir")),
            "valRetAir": xml_decimal(first_air.get("valRetAir")),
            "air_detalle": air_rows_to_text(air_rows[1:]),
        })
    elif clean_text(ret.get("Tipo Retencion")).upper() == "RENTA":
        row.update({
            "codRetAir": clean_text(ret.get("Codigo Retencion")),
            "baseImpAir": xml_decimal(ret.get("Base Imponible")),
            "porcentajeAir": xml_decimal(ret.get("Porcentaje Retencion")),
            "valRetAir": xml_decimal(ret.get("Valor Retenido")),
        })
    elif clean_text(ret.get("Tipo Retencion")).upper() == "IVA":
        field = retencion_iva_field(ret.get("Porcentaje Retencion"))
        if field:
            row[field] = xml_decimal(ret.get("Valor Retenido"))
    return row


def validate_ats_row(row, context=None):
    missing = [field for field in sorted(ATS_REQUIRED_XML_FIELDS) if not clean_text(row.get(field))]
    if missing:
        row["Exportar XML"] = "NO"
        row["Estado ATS"] = "REVISAR"
        row["Observacion"] = "Faltan campos obligatorios: " + ", ".join(missing)
        return row

    if FILTRAR_ATS_POR_PERIODO:
        context = context or row
        parsed = parse_date_any(clean_text(row.get("fechaEmision"))) or parse_date_any(clean_text(row.get("fechaRegistro")))
        if parsed:
            expected_year = int(clean_text(context.get("Anio")))
            expected_month = int(clean_text(context.get("Mes")))
            if parsed.year != expected_year or parsed.month != expected_month:
                row["Exportar XML"] = "NO"
                row["Estado ATS"] = "FUERA_PERIODO"
                row["Observacion"] = f"Documento fuera del periodo ATS configurado {expected_month:02d}/{expected_year}."
    return row


def build_ats_dataframe(compras_df, nc_df, ret_df, context=None):
    retenciones_index = build_retenciones_index(ret_df)
    rows = []

    if not compras_df.empty:
        for _, compra in compras_df.iterrows():
            rows.append(ats_row_from_compra(compra, retenciones_index, context))

    if not nc_df.empty:
        for _, nota in nc_df.iterrows():
            rows.append(ats_row_from_nota_credito(nota, retenciones_index, context))

    for info in retenciones_index.values():
        if not info["used"]:
            rows.append(ats_row_from_retencion_pendiente(info, context))

    if not rows:
        return pd.DataFrame(columns=ATS_COLUMNS)
    return pd.DataFrame(rows)[ATS_COLUMNS]


def add_xml_text(parent, name, value):
    child = ET.SubElement(parent, name)
    child.text = clean_text(value)
    return child


def add_pagos_exterior(detalle, row):
    pago = ET.SubElement(detalle, "pagoExterior")
    add_xml_text(pago, "pagoLocExt", clean_text(row.get("pagoLocExt")) or ATS_PAGO_LOC_EXT_DEFAULT)
    add_xml_text(pago, "paisEfecPago", clean_text(row.get("paisEfecPago")) or ATS_PAIS_EFEC_PAGO_DEFAULT)
    add_xml_text(pago, "aplicConvDobTrib", clean_text(row.get("aplicConvDobTrib")) or ATS_APLIC_CONV_DOB_TRIB_DEFAULT)
    add_xml_text(pago, "pagExtSujRetNorLeg", clean_text(row.get("pagExtSujRetNorLeg")) or ATS_PAG_EXT_SUJ_RET_NOR_LEG_DEFAULT)


def add_formas_pago(detalle, row):
    add_formas_pago_value(detalle, row.get("formaPago"))


def add_formas_pago_value(detalle, value):
    forma_pago = clean_text(value)
    if not forma_pago:
        return
    formas = ET.SubElement(detalle, "formasDePago")
    for item in re.split(r"[;,]", forma_pago):
        item = item.strip()
        if item:
            add_xml_text(formas, "formaPago", item)


def add_air(detalle, row):
    air_rows = []
    first_air = air_row_from_ats_columns(row)
    if first_air:
        air_rows.append(first_air)
    air_rows.extend(parse_air_text(row.get("air_detalle")))
    if not air_rows:
        return
    air = ET.SubElement(detalle, "air")
    for air_row in air_rows:
        detalle_air = ET.SubElement(air, "detalleAir")
        add_xml_text(detalle_air, "codRetAir", air_row["codRetAir"])
        add_xml_text(detalle_air, "baseImpAir", xml_decimal(air_row["baseImpAir"]))
        add_xml_text(detalle_air, "porcentajeAir", xml_decimal(air_row["porcentajeAir"]))
        add_xml_text(detalle_air, "valRetAir", xml_decimal(air_row["valRetAir"]))


def add_optional_retencion(detalle, row):
    if not clean_text(row.get("estabRetencion1")):
        return
    add_xml_text(detalle, "estabRetencion1", clean_fixed_digits(row.get("estabRetencion1"), 3))
    add_xml_text(detalle, "ptoEmiRetencion1", clean_fixed_digits(row.get("ptoEmiRetencion1"), 3))
    add_xml_text(detalle, "secRetencion1", clean_integer_text(row.get("secRetencion1")))
    if clean_text(row.get("autRetencion1")):
        add_xml_text(detalle, "autRetencion1", clean_digits(row.get("autRetencion1")))
    if clean_text(row.get("fechaEmiRet1")):
        add_xml_text(detalle, "fechaEmiRet1", date_ddmmyyyy(row.get("fechaEmiRet1")))


def add_optional_doc_modificado(detalle, row):
    if not clean_text(row.get("docModificado")):
        return
    add_xml_text(detalle, "docModificado", clean_integer_text(row.get("docModificado"), strip_leading=False).zfill(2))
    add_xml_text(detalle, "estabModificado", clean_fixed_digits(row.get("estabModificado"), 3))
    add_xml_text(detalle, "ptoEmiModificado", clean_fixed_digits(row.get("ptoEmiModificado"), 3))
    add_xml_text(detalle, "secModificado", clean_integer_text(row.get("secModificado")))
    if clean_text(row.get("autModificado")):
        add_xml_text(detalle, "autModificado", clean_digits(row.get("autModificado")))


def append_detalle_compra(parent, row):
    detalle = ET.SubElement(parent, "detalleCompras")
    for field in ATS_DETAIL_XML_ORDER:
        value = row.get(field)
        if field in ATS_OPTIONAL_XML_FIELDS and not clean_text(value):
            continue
        if field in ATS_NUMERIC_COLS:
            value = xml_decimal(value)
        elif field in ATS_DATE_COLS:
            value = date_ddmmyyyy(value)
        elif field in {"establecimiento", "puntoEmision"}:
            value = clean_fixed_digits(value, 3)
        elif field == "secuencial":
            value = clean_integer_text(value)
        elif field == "autorizacion":
            value = clean_digits(value)
        elif field == "tpIdProv":
            value = clean_fixed_digits(value, 2)
        elif field == "tipoComprobante":
            value = clean_integer_text(value, strip_leading=False).zfill(2)
        add_xml_text(detalle, field, value)

    add_pagos_exterior(detalle, row)
    add_formas_pago(detalle, row)
    add_air(detalle, row)
    add_optional_retencion(detalle, row)
    add_optional_doc_modificado(detalle, row)


def ventas_exportable_df(ventas_df):
    if ventas_df is None or ventas_df.empty:
        return pd.DataFrame(columns=VENTAS_COLUMNS)
    export_df = ventas_df.copy()
    if "Exportar XML" in export_df.columns:
        export_df = export_df[export_df["Exportar XML"].astype(str).str.upper().str.strip() != "NO"]
    if "Numero de Factura" in export_df.columns:
        export_df = export_df[export_df["Numero de Factura"].apply(clean_text) != ""]
    for field in ["tpIdCliente", "idCliente", "tipoComprobante"]:
        if field in export_df.columns:
            export_df = export_df[export_df[field].apply(clean_text) != ""]
    return export_df


def venta_base_no_gra(row):
    return to_float(row.get("No Objeto IVA"))


def venta_base_0(row):
    return to_float(row.get("Subtotal IVA 0%"))


def venta_base_gravada(row):
    return sum_columns(row, [
        "Subtotal IVA 5%",
        "Subtotal IVA 8%",
        "Subtotal IVA 12%",
        "Subtotal IVA 14%",
        "Subtotal IVA 15%",
    ])


def venta_total_bases(row):
    return venta_base_no_gra(row) + venta_base_0(row) + venta_base_gravada(row)


def ventas_total_bases(ventas_df):
    return sum(venta_total_bases(row) for _, row in ventas_exportable_df(ventas_df).iterrows())


def ventas_grouped_rows(ventas_df):
    grouped = {}
    for _, row in ventas_exportable_df(ventas_df).iterrows():
        forma_pago = clean_text(row.get("Forma Pago"))
        key = (
            clean_fixed_digits(row.get("tpIdCliente"), 2),
            clean_digits(row.get("idCliente") or row.get("Identificacion Cliente")),
            clean_text(row.get("parteRelVtas")) or "NO",
            clean_integer_text(row.get("tipoComprobante"), strip_leading=False).zfill(2),
            clean_text(row.get("tipoEmision")) or "E",
            forma_pago,
        )
        item = grouped.setdefault(key, {
            "tpIdCliente": key[0],
            "idCliente": key[1],
            "parteRelVtas": key[2],
            "tipoComprobante": key[3],
            "tipoEmision": key[4],
            "numeroComprobantes": 0,
            "baseNoGraIva": 0.0,
            "baseImponible": 0.0,
            "baseImpGrav": 0.0,
            "montoIva": 0.0,
            "montoIce": 0.0,
            "valorRetIva": 0.0,
            "valorRetRenta": 0.0,
            "formaPago": forma_pago,
        })
        item["numeroComprobantes"] += int(to_float(row.get("numeroComprobantes") or 1) or 1)
        item["baseNoGraIva"] += venta_base_no_gra(row)
        item["baseImponible"] += venta_base_0(row)
        item["baseImpGrav"] += venta_base_gravada(row)
        item["montoIva"] += sum_columns(row, ["IVA 5%", "IVA 8%", "IVA 12%", "IVA 14%", "IVA 15%"])
        item["montoIce"] += to_float(row.get("ICE"))
        item["valorRetIva"] += to_float(row.get("valorRetIva"))
        item["valorRetRenta"] += to_float(row.get("valorRetRenta"))
    return list(grouped.values())


def append_detalle_venta(parent, row):
    detalle = ET.SubElement(parent, "detalleVentas")
    add_xml_text(detalle, "tpIdCliente", clean_fixed_digits(row.get("tpIdCliente"), 2))
    add_xml_text(detalle, "idCliente", clean_digits(row.get("idCliente")))
    if clean_text(row.get("parteRelVtas")):
        add_xml_text(detalle, "parteRelVtas", clean_text(row.get("parteRelVtas")))
    add_xml_text(detalle, "tipoComprobante", clean_text(row.get("tipoComprobante")))
    add_xml_text(detalle, "tipoEmision", clean_text(row.get("tipoEmision")) or "E")
    add_xml_text(detalle, "numeroComprobantes", clean_integer_text(row.get("numeroComprobantes")))
    add_xml_text(detalle, "baseNoGraIva", xml_decimal(row.get("baseNoGraIva")))
    add_xml_text(detalle, "baseImponible", xml_decimal(row.get("baseImponible")))
    add_xml_text(detalle, "baseImpGrav", xml_decimal(row.get("baseImpGrav")))
    add_xml_text(detalle, "montoIva", xml_decimal(row.get("montoIva")))
    add_xml_text(detalle, "montoIce", xml_decimal(row.get("montoIce")))
    add_xml_text(detalle, "valorRetIva", xml_decimal(row.get("valorRetIva")))
    add_xml_text(detalle, "valorRetRenta", xml_decimal(row.get("valorRetRenta")))
    add_formas_pago_value(detalle, row.get("formaPago"))


def add_ventas_to_xml(root, ventas_df):
    rows = ventas_grouped_rows(ventas_df)
    if not rows:
        return
    ventas = ET.SubElement(root, "ventas")
    for row in rows:
        append_detalle_venta(ventas, row)


def add_ventas_establecimiento_to_xml(root, ventas_df):
    export_df = ventas_exportable_df(ventas_df)
    if export_df.empty:
        return
    totals = {}
    for _, row in export_df.iterrows():
        establecimiento = clean_fixed_digits(row.get("Establecimiento"), 3)
        if not establecimiento or establecimiento == "000":
            continue
        totals[establecimiento] = totals.get(establecimiento, 0.0) + venta_total_bases(row)
    if not totals:
        return

    ventas_est = ET.SubElement(root, "ventasEstablecimiento")
    for establecimiento, total in sorted(totals.items()):
        venta_est = ET.SubElement(ventas_est, "ventaEst")
        add_xml_text(venta_est, "codEstab", establecimiento)
        add_xml_text(venta_est, "ventasEstab", xml_decimal(total))


def build_ats_xml_tree(ats_df, ventas_df=None):
    export_df = ats_df.copy()
    if not export_df.empty and "Exportar XML" in export_df.columns:
        export_df = export_df[export_df["Exportar XML"].astype(str).str.upper().str.strip() != "NO"]

    if not export_df.empty:
        first = export_df.iloc[0]
    elif ats_df is not None and not ats_df.empty:
        first = ats_df.iloc[0]
    else:
        first = pd.Series(infer_ats_context(ventas_df=ventas_df))
    root = ET.Element("iva")
    add_xml_text(root, "TipoIDInformante", clean_text(first.get("TipoIDInformante")) or ATS_TIPO_ID_INFORMANTE)
    add_xml_text(root, "IdInformante", clean_digits(first.get("IdInformante")) or ATS_ID_INFORMANTE)
    add_xml_text(root, "razonSocial", clean_text(first.get("razonSocial")) or ATS_RAZON_SOCIAL)
    add_xml_text(root, "Anio", clean_integer_text(first.get("Anio"), strip_leading=False) or ATS_ANIO)
    add_xml_text(root, "Mes", clean_fixed_digits(first.get("Mes"), 2) or ATS_MES)
    if clean_text(first.get("regimenMicroempresa")):
        add_xml_text(root, "regimenMicroempresa", clean_text(first.get("regimenMicroempresa")))
    if clean_text(first.get("numEstabRuc")):
        add_xml_text(root, "numEstabRuc", clean_fixed_digits(first.get("numEstabRuc"), 3))
    total_ventas = to_float(first.get("totalVentas") or ATS_TOTAL_VENTAS)
    total_ventas_pdf = ventas_total_bases(ventas_df)
    if total_ventas == 0 and total_ventas_pdf:
        total_ventas = total_ventas_pdf
    add_xml_text(root, "totalVentas", xml_decimal(total_ventas))
    add_xml_text(root, "codigoOperativo", clean_text(first.get("codigoOperativo")) or ATS_CODIGO_OPERATIVO)

    if not export_df.empty:
        compras = ET.SubElement(root, "compras")
        for _, row in export_df.iterrows():
            append_detalle_compra(compras, row)

    add_ventas_to_xml(root, ventas_df)
    add_ventas_establecimiento_to_xml(root, ventas_df)

    return root


def write_ats_xml(ats_df, output_path, ventas_df=None):
    root = build_ats_xml_tree(ats_df, ventas_df)
    xml_body = ET.tostring(root, encoding="utf-8", short_empty_elements=True)
    xml_header = b'<?xml version="1.0" encoding="UTF-8" standalone="no"?>'
    Path(output_path).write_bytes(xml_header + xml_body)


def export_ats_xml_from_excel(excel_path, output_path):
    with pd.ExcelFile(excel_path) as xls:
        ats_df = pd.read_excel(xls, sheet_name="ATS", dtype=str).fillna("")
        ventas_df = (
            pd.read_excel(xls, sheet_name="Ventas", dtype=str).fillna("")
            if "Ventas" in xls.sheet_names
            else pd.DataFrame(columns=VENTAS_COLUMNS)
        )
    write_ats_xml(ats_df, output_path, ventas_df)


def ats_xml_file_path_from_excel(excel_path):
    with pd.ExcelFile(excel_path) as xls:
        if "ATS" in xls.sheet_names:
            ats_df = pd.read_excel(xls, sheet_name="ATS", dtype=str).fillna("")
            if not ats_df.empty:
                return ats_xml_file_path(ats_df.iloc[0].to_dict())
        ventas_df = (
            pd.read_excel(xls, sheet_name="Ventas", dtype=str).fillna("")
            if "Ventas" in xls.sheet_names
            else pd.DataFrame(columns=VENTAS_COLUMNS)
        )
    return ats_xml_file_path(infer_ats_context(ventas_df=ventas_df))


def style_sheet(ws, table_name, numeric_cols, date_cols, header_color, tab_color, table_style):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = tab_color

    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 36
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
            cell.border = border
            if header in numeric_cols:
                if not DECIMALES_CON_COMA and cell.value not in (None, ""):
                    cell.value = to_float(cell.value)
                cell.number_format = "@" if DECIMALES_CON_COMA else "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif header in date_cols:
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=("Razon" in str(header) or "Producto" in str(header) or "Detalle" in str(header)))

    for column_cells in ws.columns:
        max_len = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
        header = column_cells[0].value or ""
        if header in numeric_cols:
            width = 14
        elif "Razon" in str(header) or "Producto" in str(header) or "Detalle" in str(header):
            width = min(max(max_len + 2, 24), 48)
        else:
            width = min(max(max_len + 2, 12), 34)
        ws.column_dimensions[column_cells[0].column_letter].width = width

    if ws.max_row >= 2:
        ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name=table_style, showRowStripes=True)
        ws.add_table(table)


def sumproduct_formula(id_range, value_range, digits):
    pattern = "?" * digits
    if not DECIMALES_CON_COMA:
        return f'=SUMIF({id_range},"{pattern}",{value_range})'
    return f'=SUMPRODUCT((LEN({id_range})={digits})*IFERROR(NUMBERVALUE({value_range},",","."),0))'


def count_formula(id_range, digits):
    pattern = "?" * digits
    return f'=COUNTIF({id_range},"{pattern}")'


def text_number_sum_formula(value_range):
    if not DECIMALES_CON_COMA:
        return f"=SUM({value_range})"
    return f'=SUMPRODUCT(IFERROR(NUMBERVALUE({value_range},",","."),0))'


def add_identification_summary(ws, title, numeric_cols, id_col_name, header_color, subheader_color):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if ws.max_row < 2:
        return

    headers = [cell.value for cell in ws[1]]
    col_index = {name: idx + 1 for idx, name in enumerate(headers) if name}
    if id_col_name not in col_index:
        return

    numeric_cols = [column for column in numeric_cols if column in col_index]
    id_letter = ws.cell(row=1, column=col_index[id_col_name]).column_letter
    last_row = ws.max_row
    id_range = f"${id_letter}$2:${id_letter}${last_row}"
    start_row = last_row + 3

    header_fill = PatternFill("solid", fgColor=header_color)
    subheader_fill = PatternFill("solid", fgColor=subheader_color)
    white_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for label, start_col, digits in (("CEDULA", 1, 10), ("RUC", 4, 13)):
        ws.cell(start_row, start_col, f"RESUMEN {title} - {label}")
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
        ws.cell(start_row, start_col).fill = header_fill
        ws.cell(start_row, start_col).font = white_font
        ws.cell(start_row, start_col).alignment = Alignment(horizontal="center")

        ws.cell(start_row + 1, start_col, "Concepto")
        ws.cell(start_row + 1, start_col + 1, "Total")
        for cell in (ws.cell(start_row + 1, start_col), ws.cell(start_row + 1, start_col + 1)):
            cell.fill = subheader_fill
            cell.font = bold_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        first_row = start_row + 2
        ws.cell(first_row, start_col, "Cantidad comprobantes")
        ws.cell(first_row, start_col + 1, count_formula(id_range, digits))

        for offset, numeric_col in enumerate(numeric_cols, start=1):
            row = first_row + offset
            value_letter = ws.cell(row=1, column=col_index[numeric_col]).column_letter
            value_range = f"${value_letter}$2:${value_letter}${last_row}"
            ws.cell(row, start_col, numeric_col)
            ws.cell(row, start_col + 1, sumproduct_formula(id_range, value_range, digits))

        for row in range(first_row, first_row + len(numeric_cols) + 1):
            ws.cell(row, start_col).border = border
            ws.cell(row, start_col + 1).border = border
            ws.cell(row, start_col + 1).number_format = "0" if row == first_row else "#,##0.00"
            ws.cell(row, start_col + 1).alignment = Alignment(horizontal="right")

        ws.column_dimensions[ws.cell(1, start_col).column_letter].width = max(ws.column_dimensions[ws.cell(1, start_col).column_letter].width or 0, 28)
        ws.column_dimensions[ws.cell(1, start_col + 1).column_letter].width = max(ws.column_dimensions[ws.cell(1, start_col + 1).column_letter].width or 0, 16)


def add_retenciones_summary(ws):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if ws.max_row < 2:
        return

    headers = [cell.value for cell in ws[1]]
    col_index = {name: idx + 1 for idx, name in enumerate(headers) if name}
    if "Tipo Retencion" not in col_index or "Valor Retenido" not in col_index:
        return

    tipo_letter = ws.cell(1, col_index["Tipo Retencion"]).column_letter
    valor_letter = ws.cell(1, col_index["Valor Retenido"]).column_letter
    last_row = ws.max_row
    start_row = last_row + 3

    header_fill = PatternFill("solid", fgColor="548235")
    subheader_fill = PatternFill("solid", fgColor="E2F0D9")
    white_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="A9D18E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(start_row, 1, "RESUMEN RETENCIONES")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    ws.cell(start_row, 1).fill = header_fill
    ws.cell(start_row, 1).font = white_font
    ws.cell(start_row, 1).alignment = Alignment(horizontal="center")

    ws.cell(start_row + 1, 1, "Tipo Retencion")
    ws.cell(start_row + 1, 2, "Valor Retenido")
    for cell in (ws.cell(start_row + 1, 1), ws.cell(start_row + 1, 2)):
        cell.fill = subheader_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    first_row = start_row + 2
    for offset, tipo in enumerate(("IVA", "RENTA", "ISD")):
        row = first_row + offset
        ws.cell(row, 1, tipo)
        if not DECIMALES_CON_COMA:
            formula = f'=SUMIF(${tipo_letter}$2:${tipo_letter}${last_row},"{tipo}",${valor_letter}$2:${valor_letter}${last_row})'
        else:
            formula = f'=SUMPRODUCT((${tipo_letter}$2:${tipo_letter}${last_row}="{tipo}")*IFERROR(NUMBERVALUE(${valor_letter}$2:${valor_letter}${last_row},",","."),0))'
        ws.cell(row, 2, formula)

    total_row = first_row + 3
    ws.cell(total_row, 1, "TOTAL")
    ws.cell(total_row, 2, f"=SUM(B{first_row}:B{total_row - 1})")

    for row in range(first_row, total_row + 1):
        ws.cell(row, 1).border = border
        ws.cell(row, 2).border = border
        ws.cell(row, 2).number_format = "#,##0.00"
        ws.cell(row, 2).alignment = Alignment(horizontal="right")
    ws.cell(total_row, 1).font = bold_font
    ws.cell(total_row, 2).font = bold_font


def add_ventas_summary(ws):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if ws.max_row < 2:
        return

    headers = [cell.value for cell in ws[1]]
    col_index = {name: idx + 1 for idx, name in enumerate(headers) if name}
    if "Numero de Factura" not in col_index:
        return

    last_row = ws.max_row
    start_row = last_row + 3

    header_fill = PatternFill("solid", fgColor="0F6B78")
    subheader_fill = PatternFill("solid", fgColor="DDEBF7")
    white_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="9EADCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(start_row, 1, "RESUMEN VENTAS")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    ws.cell(start_row, 1).fill = header_fill
    ws.cell(start_row, 1).font = white_font
    ws.cell(start_row, 1).alignment = Alignment(horizontal="center")

    ws.cell(start_row + 1, 1, "Concepto")
    ws.cell(start_row + 1, 2, "Total")
    for cell in (ws.cell(start_row + 1, 1), ws.cell(start_row + 1, 2)):
        cell.fill = subheader_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    factura_letter = ws.cell(1, col_index["Numero de Factura"]).column_letter
    rows = [("Cantidad comprobantes", f'=COUNTA(${factura_letter}$2:${factura_letter}${last_row})')]

    for label in VENTAS_SUMMARY_COLS:
        if label in col_index:
            value_letter = ws.cell(1, col_index[label]).column_letter
            rows.append((label, text_number_sum_formula(f"${value_letter}$2:${value_letter}${last_row}")))

    first_row = start_row + 2
    for offset, (label, formula) in enumerate(rows):
        row = first_row + offset
        ws.cell(row, 1, label)
        ws.cell(row, 2, formula)
        ws.cell(row, 1).border = border
        ws.cell(row, 2).border = border
        ws.cell(row, 2).alignment = Alignment(horizontal="right")
        ws.cell(row, 2).number_format = "0" if offset == 0 else "#,##0.00"

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 28)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 16)


def catalog_code_range(column_letter, rows, min_last_row=3):
    last_row = max(min_last_row, len(rows) + 2)
    return f"'Catalogos ATS'!${column_letter}$3:${column_letter}${last_row}"


def clean_catalog_text(value):
    return clean_pdf_cell(value).replace("\n", " ").strip()


def load_retencion_renta_catalog():
    global _RETENCION_RENTA_CATALOG_CACHE
    if _RETENCION_RENTA_CATALOG_CACHE is not None:
        return _RETENCION_RENTA_CATALOG_CACHE

    ficha_path = Path(ATS_FICHA_TECNICA_PDF_MANUAL)
    rows = []
    if not ficha_path.exists():
        _RETENCION_RENTA_CATALOG_CACHE = rows
        return rows

    try:
        import pdfplumber
    except Exception:
        _RETENCION_RENTA_CATALOG_CACHE = rows
        return rows

    seen = set()
    try:
        with pdfplumber.open(ficha_path) as pdf:
            for page_no in range(72, 79):
                if page_no - 1 >= len(pdf.pages):
                    continue
                for table in pdf.pages[page_no - 1].extract_tables() or []:
                    for table_row in table:
                        if len(table_row) < 4:
                            continue
                        modulo = clean_catalog_text(table_row[0])
                        codigo = clean_catalog_text(table_row[1]).replace(" ", "")
                        concepto = clean_catalog_text(table_row[2])
                        if not re.match(r"^\d{3,4}[A-Za-z0-9]*$", codigo or ""):
                            continue
                        if not concepto or codigo in seen:
                            continue
                        porcentaje = ""
                        for candidate in table_row[3:]:
                            text = clean_catalog_text(candidate)
                            if text:
                                porcentaje = text
                                break
                        rows.append((codigo, modulo, concepto, porcentaje))
                        seen.add(codigo)
    except Exception:
        rows = []

    _RETENCION_RENTA_CATALOG_CACHE = rows
    return rows


def style_ats_sheet(ws):
    from openpyxl.comments import Comment
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    style_sheet(
        ws,
        "TablaATS",
        ATS_NUMERIC_COLS,
        ATS_DATE_COLS,
        "C65911",
        "C65911",
        "TableStyleMedium3",
    )

    headers = [cell.value for cell in ws[1]]
    editable_header_fill = PatternFill("solid", fgColor="C00000")
    editable_cell_fill = PatternFill("solid", fgColor="FFE5E5")
    editable_header_font = Font(bold=True, color="FFFFFF")
    editable_side = Side(style="medium", color="C00000")
    editable_border = Border(left=editable_side, right=editable_side)

    for col_idx, header in enumerate(headers, start=1):
        if header in ATS_CODE_TEXT_COLS:
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                for item in cell:
                    item.number_format = "@"
        if header in ATS_EDITABLE_COLUMNS:
            ws.cell(1, col_idx).fill = editable_header_fill
            ws.cell(1, col_idx).font = editable_header_font
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).fill = editable_cell_fill
                ws.cell(row, col_idx).border = editable_border

    def column_letter(header):
        if header not in headers:
            return None
        return ws.cell(1, headers.index(header) + 1).column_letter

    def add_dropdown(header, values=None, prompt="", formula1=None):
        letter = column_letter(header)
        if not letter:
            return
        last_row = max(ws.max_row + 200, 500)
        formula = formula1 or ('"' + ",".join(values) + '"')
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Selecciona un codigo del catalogo ATS."
        dv.errorTitle = "Codigo ATS no valido"
        dv.prompt = prompt
        dv.promptTitle = header
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last_row}")

    retencion_renta_catalog = load_retencion_renta_catalog()
    add_dropdown("codSustento", prompt="Tabla 5. 01 IVA credito tributario; 02 costo/gasto IR; 03/04 activos fijos; 06/07 inventario; 08 reembolso; 15 servicios digitales.", formula1=catalog_code_range("A", ATS_SUSTENTO_CATALOG))
    add_dropdown("tipoComprobante", prompt="Tabla 4. 01 factura; 02 nota venta; 03 liquidacion; 04 nota credito; 05 nota debito; 18 ventas; 41 reembolso; hasta 375 segun catalogo.", formula1=catalog_code_range("F", ATS_TIPO_COMPROBANTE_CATALOG))
    add_dropdown("parteRel", ["NO", "SI"], "SI solo si el proveedor/cliente es parte relacionada.")
    add_dropdown("tpIdProv", ["01", "02", "03"], "01 RUC; 02 cedula; 03 pasaporte/otros.")
    add_dropdown("pagoLocExt", ["01", "02"], "01 pago local/residente; 02 pago al exterior/no residente.")
    add_dropdown("paisEfecPago", prompt="NA para pago local. Para exterior, usar codigo de pais de Catalogos ATS.", formula1="'Catalogos ATS'!$R$3:$R$250")
    add_dropdown("aplicConvDobTrib", ["NA", "SI", "NO"], "NA pago local; SI/NO si aplica convenio de doble tributacion en pago exterior.")
    add_dropdown("pagExtSujRetNorLeg", ["NA", "SI", "NO"], "NA pago local; SI/NO si el pago exterior esta sujeto a retencion segun norma legal.")
    add_dropdown("formaPago", prompt="Tabla 13. Vigentes comunes: 01, 15, 16, 17, 18, 19, 20, 21. El catalogo tambien muestra 02-14 como vencidos.", formula1=catalog_code_range("J", ATS_FORMA_PAGO_CATALOG))
    add_dropdown("codRetAir", prompt="Tabla 3.10 AIR. Ejemplos: 303 honorarios, 304/304A servicios, 332 sin retencion, 343 otras 1%, 3440 otras 2.75%.", formula1=catalog_code_range("W", retencion_renta_catalog, min_last_row=120))
    add_dropdown("docModificado", prompt="Solo para notas de credito/debito: tipo del documento modificado. Normalmente 01 si modifica factura.", formula1=catalog_code_range("F", ATS_TIPO_COMPROBANTE_CATALOG))

    header_comments = {
        "codSustento": "Manual critico. 01 credito tributario IVA; 02 costo/gasto IR; 03/04 activos fijos; 06/07 inventario; 08 reembolso; 15 servicios digitales.",
        "tipoComprobante": "Manual critico. Compras: 01 factura, 02 nota venta, 03 liquidacion, 04 nota credito, 05 nota debito. Ventas ATS usa 18. Revisa Catalogos ATS para codigos hasta 375.",
        "pagoLocExt": "01 pago local/residente. 02 pago al exterior/no residente.",
        "paisEfecPago": "NA para pagos locales. Si pagoLocExt es 02, usar el codigo de pais de la hoja Catalogos ATS.",
        "aplicConvDobTrib": "NA para pago local. Para exterior, indicar SI o NO.",
        "pagExtSujRetNorLeg": "NA para pago local. Para exterior, indicar SI o NO segun aplique retencion legal.",
        "formaPago": "Tabla 13. Vigentes comunes: 01 sin sistema financiero; 15 compensacion; 16 debito; 17 dinero electronico; 18 prepago; 19 credito; 20 otros sistema financiero; 21 endoso.",
        "codRetAir": "Retencion en la fuente de Impuesto a la Renta. Se toma del XML de retencion si existe; si se edita manualmente, usa la Tabla 3.10 en Catalogos ATS.",
        "baseImpAir": "Base imponible de la retencion AIR asociada al codigo codRetAir.",
        "porcentajeAir": "Porcentaje aplicado al codigo AIR. Debe coincidir con el catalogo y el periodo.",
        "valRetAir": "Valor retenido de renta. Si hay varias retenciones AIR en un comprobante, usa air_detalle para las adicionales.",
        "air_detalle": "Retenciones AIR adicionales en formato codigo|base|porcentaje|valor separadas por punto y coma. Ejemplo: 303|100.00|10.00|10.00",
        "docModificado": "Solo notas de credito/debito. Normalmente 01 cuando la nota modifica una factura.",
    }
    for header, body in header_comments.items():
        letter = column_letter(header)
        if letter:
            ws[f"{letter}1"].comment = Comment(body, "Codex")

    status_col = headers.index("Estado ATS") + 1 if "Estado ATS" in headers else None
    if status_col:
        revisar_fill = PatternFill("solid", fgColor="FCE4D6")
        for row in range(2, ws.max_row + 1):
            if clean_text(ws.cell(row, status_col).value).upper() not in {"", "OK"}:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = revisar_fill


def style_ventas_sheet(ws):
    from openpyxl.styles import PatternFill

    style_sheet(
        ws,
        "TablaVentas",
        VENTAS_NUMERIC_COLS,
        VENTAS_DATE_COLS,
        "0F6B78",
        "0F6B78",
        "TableStyleMedium4",
    )

    headers = [cell.value for cell in ws[1]]
    for col_idx, header in enumerate(headers, start=1):
        if header in VENTAS_CODE_TEXT_COLS:
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                for item in cell:
                    item.number_format = "@"

    status_col = headers.index("Estado ATS") + 1 if "Estado ATS" in headers else None
    if status_col:
        revisar_fill = PatternFill("solid", fgColor="FCE4D6")
        for row in range(2, ws.max_row + 1):
            if clean_text(ws.cell(row, status_col).value).upper() not in {"", "OK"}:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = revisar_fill


def load_paises_catalog():
    ficha_path = Path(ATS_FICHA_TECNICA_PDF_MANUAL)
    rows = [("NA", "No aplica / pago local")]
    if not ficha_path.exists():
        return rows

    try:
        import pdfplumber
    except Exception:
        return rows

    found = {}
    try:
        with pdfplumber.open(ficha_path) as pdf:
            for page_no in (84, 85, 86):
                if page_no - 1 >= len(pdf.pages):
                    continue
                for table in pdf.pages[page_no - 1].extract_tables() or []:
                    title = normalize_header(table[0][0] if table and table[0] else "")
                    if not table or "tabla 16" not in title or "pais" not in title:
                        continue
                    for row in table[2:]:
                        for idx in (0, 2, 4):
                            if idx + 1 >= len(row):
                                continue
                            pais = clean_pdf_cell(row[idx])
                            codigo = clean_digits(row[idx + 1])
                            if pais and codigo:
                                found[codigo.zfill(3)] = pais
    except Exception:
        return rows

    for codigo in sorted(found, key=lambda item: int(item)):
        rows.append((codigo, found[codigo]))
    return rows


def add_catalogos_ats_sheet(wb):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if "Catalogos ATS" in wb.sheetnames:
        del wb["Catalogos ATS"]

    ws = wb.create_sheet("Catalogos ATS")
    ws.sheet_properties.tabColor = "FFC000"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    white_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    paises_catalog = load_paises_catalog()
    retencion_renta_catalog = load_retencion_renta_catalog()
    sections = [
        ("A", "SUSTENTO TRIBUTARIO", ["Codigo", "Tipo de sustento", "Codigos tipo comprobante permitidos"], ATS_SUSTENTO_CATALOG),
        ("F", "TIPO DE COMPROBANTE", ["Codigo", "Tipo de comprobante"], ATS_TIPO_COMPROBANTE_CATALOG),
        ("J", "FORMAS DE PAGO / COBRO", ["Codigo", "Forma", "Fecha Inicio", "Fecha Fin"], ATS_FORMA_PAGO_CATALOG),
        ("O", "PAGO LOCAL / EXTERIOR", ["Codigo", "Descripcion"], ATS_TIPO_PAGO_CATALOG),
        ("R", "PAISES (TABLA 16)", ["Codigo", "Pais"], paises_catalog),
        ("U", "SI / NO / NA", ["Codigo", "Descripcion"], ATS_SI_NO_NA_CATALOG),
        ("W", "RETENCION RENTA AIR (TABLA 3.10)", ["Codigo", "Modulo", "Concepto", "% vigente"], retencion_renta_catalog),
        ("AB", "RETENCION IVA (TABLA 11)", ["Codigo", "% IVA", "Fecha Inicio", "Fecha Fin"], ATS_RETENCION_IVA_CATALOG),
    ]

    for start_col_letter, title, headers, rows in sections:
        start_col = ws[start_col_letter + "1"].column
        end_col = start_col + len(headers) - 1
        ws.cell(1, start_col, title)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(1, start_col).fill = title_fill
        ws.cell(1, start_col).font = white_font
        ws.cell(1, start_col).alignment = Alignment(horizontal="center")

        for offset, header in enumerate(headers):
            cell = ws.cell(2, start_col + offset, header)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row_idx, row_values in enumerate(rows, start=3):
            for offset, value in enumerate(row_values):
                cell = ws.cell(row_idx, start_col + offset, value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if offset == 0:
                    cell.fill = editable_fill
                    cell.number_format = "@"

    widths = {
        "A": 12, "B": 58, "C": 42,
        "F": 12, "G": 58,
        "J": 12, "K": 46, "L": 14, "M": 14,
        "O": 12, "P": 42,
        "R": 12, "S": 36,
        "U": 12, "V": 22,
        "W": 12, "X": 20, "Y": 74, "Z": 24,
        "AB": 12, "AC": 12, "AD": 14, "AE": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def format_workbook(path):
    from openpyxl import load_workbook

    wb = load_workbook(path)

    style_sheet(wb["Compras"], "TablaCompras", COMPRAS_NUMERIC_COLS, {"Fecha Emision", "Fecha Autorizacion"}, "1F4E78", "1F4E78", "TableStyleMedium2")
    add_identification_summary(wb["Compras"], "COMPRAS", COMPRAS_NUMERIC_COLS, "Identificacion Comprador", "1F4E78", "D9EAF7")

    style_sheet(wb["NotasCredito"], "TablaNotasCredito", NC_NUMERIC_COLS, {"Fecha Emision"}, "7030A0", "7030A0", "TableStyleMedium5")
    add_identification_summary(wb["NotasCredito"], "NOTAS CREDITO", NC_NUMERIC_COLS, "Identificacion Comprador", "7030A0", "EADCF8")

    style_sheet(wb["Retenciones"], "TablaRetenciones", RET_NUMERIC_COLS, {"Fecha Emision", "Fecha Doc Sustento"}, "548235", "548235", "TableStyleMedium4")
    add_retenciones_summary(wb["Retenciones"])

    style_ventas_sheet(wb["Ventas"])
    add_ventas_summary(wb["Ventas"])

    if "RetencionesEmitidas" in wb.sheetnames:
        style_sheet(wb["RetencionesEmitidas"], "TablaRetencionesEmitidas", RET_EMITIDAS_NUMERIC_COLS, RET_EMITIDAS_DATE_COLS, "0F6B78", "0F6B78", "TableStyleMedium6")
        add_retenciones_summary(wb["RetencionesEmitidas"])

    if "Alertas" in wb.sheetnames:
        style_sheet(wb["Alertas"], "TablaAlertas", [], set(), "C00000", "C00000", "TableStyleMedium3")

    add_catalogos_ats_sheet(wb)
    style_ats_sheet(wb["ATS"])

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(path)


def process_xml_files():
    compras_rows = []
    nc_rows = []
    ret_rows = []
    ret_doc_uids = set()
    skipped_unknown = 0
    errors = []
    processed_xml_files = []
    error_xml_files = []

    for xml_file in list_xml_files(XML_FOLDER, RECURSIVE):
        try:
            root, numero_aut, fecha_aut = unwrap_xml(xml_file)
            tag = (root.tag or "").lower()

            if tag == "factura":
                compras_rows.append(parse_compra_row(root, numero_aut, fecha_aut))
                processed_xml_files.append(xml_file)
            elif tag == "notacredito":
                nc_rows.append(parse_nc_row(root, numero_aut))
                processed_xml_files.append(xml_file)
            elif tag == "comprobanteretencion":
                rows, doc_uid = parse_retencion_rows(root, numero_aut)
                if doc_uid and doc_uid in ret_doc_uids:
                    processed_xml_files.append(xml_file)
                    continue
                if doc_uid:
                    ret_doc_uids.add(doc_uid)
                ret_rows.extend(rows)
                processed_xml_files.append(xml_file)
            else:
                skipped_unknown += 1
                error_xml_files.append(xml_file)
        except Exception as exc:
            errors.append(f"{xml_file.name}: {exc}")
            error_xml_files.append(xml_file)

    compras_df, compras_dups = prepare_dataframe(
        compras_rows,
        COMPRAS_COLUMNS,
        COMPRAS_NUMERIC_COLS,
        ["RUC Emisor", "Numero de Factura", "Fecha Emision"],
    )
    nc_df, nc_dups = prepare_dataframe(
        nc_rows,
        NC_COLUMNS,
        NC_NUMERIC_COLS,
        ["RUC Emisor", "Numero de Nota de Crédito", "Fecha Emision"],
    )
    ret_df, ret_dups = prepare_dataframe(
        ret_rows,
        RET_COLUMNS,
        RET_NUMERIC_COLS,
        [
            "Numero Autorizacion",
            "Clave Acceso",
            "Numero Doc Sustento",
            "Tipo Retencion",
            "Codigo Retencion",
            "Base Imponible",
            "Porcentaje Retencion",
            "Valor Retenido",
        ],
    )
    ventas_df, ret_emitidas_df, pdf_errors = process_pdf_ventas_y_retenciones_emitidas()
    ats_context = infer_ats_context(compras_df, nc_df, ventas_df)
    ventas_df = validate_ventas_dataframe(ventas_df, ats_context)
    contribuyente = infer_contribuyente(compras_df, nc_df, ret_df, ventas_df)
    alertas_df = build_alertas_dataframe(compras_df, nc_df, ret_df, ventas_df, ret_emitidas_df, contribuyente)
    ats_df = build_ats_dataframe(compras_df, nc_df, ret_df, ats_context)

    return {
        "compras_df": compras_df,
        "nc_df": nc_df,
        "ret_df": ret_df,
        "ventas_df": ventas_df,
        "ret_emitidas_df": ret_emitidas_df,
        "ats_df": ats_df,
        "ats_context": ats_context,
        "contribuyente": contribuyente,
        "alertas_df": alertas_df,
        "compras_dups": compras_dups,
        "nc_dups": nc_dups,
        "ret_dups": ret_dups,
        "skipped_unknown": skipped_unknown,
        "errors": errors + [f"PDF ventas - {error}" for error in pdf_errors],
        "processed_xml_files": processed_xml_files,
        "error_xml_files": error_xml_files,
    }


def next_available_path(path):
    if not path.exists():
        return path
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_{timestamp}{path.suffix}")


def move_xml_files(files, destination):
    destination.mkdir(parents=True, exist_ok=True)
    moved = 0
    for source in files:
        source = Path(source)
        if not source.exists():
            continue
        target = next_available_path(destination / source.name)
        shutil.move(str(source), str(target))
        moved += 1
    return moved


def main():
    base_folders = [XML_FOLDER, SALIDA_EXCEL_DIR]
    if MOVER_XML_PROCESADOS:
        base_folders.extend([PROCESADOS_DIR, ERRORES_DIR])
    for folder in base_folders:
        folder.mkdir(parents=True, exist_ok=True)

    verificar_activacion()

    for folder in (
        PDF_ROOT_DIR,
        PDF_COMPRAS_FOLDER,
        PDF_NC_RECIBIDAS_FOLDER,
        PDF_RET_RECIBIDAS_FOLDER,
        PDF_VENTAS_FOLDER,
    ):
        folder.mkdir(parents=True, exist_ok=True)

    result = process_xml_files()
    output_file = excel_output_path(result["contribuyente"])
    output_file.parent.mkdir(parents=True, exist_ok=True)

    alertas_df = result["alertas_df"]
    ret_emitidas_con_alerta = False
    if not alertas_df.empty:
        for _, alerta in alertas_df.iterrows():
            tipo = clean_text(alerta.get("Tipo"))
            detalle = clean_text(alerta.get("Detalle"))
            if tipo == "NO COINCIDE EMISOR" or "Retenciones emitidas" in detalle:
                ret_emitidas_con_alerta = True
                break

    if ret_emitidas_con_alerta and not result["ret_emitidas_df"].empty:
        result["ret_emitidas_df"] = result["ret_emitidas_df"].copy()
        result["ret_emitidas_df"]["Estado"] = "REVISAR"
        result["ret_emitidas_df"]["Observacion"] = "No coincide con el contribuyente principal. Revisar hoja Alertas."

    try:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            result["compras_df"].to_excel(writer, index=False, sheet_name="Compras")
            result["nc_df"].to_excel(writer, index=False, sheet_name="NotasCredito")
            result["ret_df"].to_excel(writer, index=False, sheet_name="Retenciones")
            result["ventas_df"].to_excel(writer, index=False, sheet_name="Ventas")
            result["ret_emitidas_df"].to_excel(writer, index=False, sheet_name="RetencionesEmitidas")
            result["alertas_df"].to_excel(writer, index=False, sheet_name="Alertas")
            result["ats_df"].to_excel(writer, index=False, sheet_name="ATS")

        format_workbook(output_file)
    except PermissionError:
        print(f"No pude guardar porque el Excel esta abierto o bloqueado: {output_file}")
        print("Cierra ese archivo de Excel y vuelve a ejecutar.")
        return

    moved_ok = 0
    moved_errors = 0
    if MOVER_XML_PROCESADOS:
        moved_ok = move_xml_files(result["processed_xml_files"], PROCESADOS_DIR)
        moved_errors = move_xml_files(result["error_xml_files"], ERRORES_DIR)

    print(f"Listo: {output_file}")
    print(f"Entrada XML: {XML_FOLDER}")
    print(f"Salida Excel: {SALIDA_EXCEL_DIR}")
    print(f"Compras: {len(result['compras_df'])} filas | duplicados omitidos: {result['compras_dups']}")
    print(f"Notas de crédito: {len(result['nc_df'])} filas | duplicados omitidos: {result['nc_dups']}")
    print(f"Retenciones: {len(result['ret_df'])} filas | duplicados omitidos: {result['ret_dups']}")
    print(f"Ventas PDF: {len(result['ventas_df'])} filas")
    print(f"Retenciones emitidas PDF: {len(result['ret_emitidas_df'])} filas")
    if ret_emitidas_con_alerta and not result["ret_emitidas_df"].empty:
        print("ALERTA: las retenciones emitidas no concuerdan con el contribuyente principal. Revisa la hoja Alertas.")
    print(f"ATS: {len(result['ats_df'])} filas en la hoja ATS")
    if MOVER_XML_PROCESADOS:
        print(f"XML movidos a procesados: {moved_ok}")
        print(f"XML movidos a errores: {moved_errors}")
    if GENERAR_XML_ATS:
        ats_xml_file = ats_xml_file_path(result["ats_context"])
        export_ats_xml_from_excel(output_file, ats_xml_file)
        print(f"XML ATS: {ats_xml_file}")
    if result["skipped_unknown"]:
        print(f"XML no reconocidos: {result['skipped_unknown']}")
    if result["errors"]:
        print("Errores:")
        for error in result["errors"]:
            print(f"  - {error}")


if __name__ == "__main__":
    main()
